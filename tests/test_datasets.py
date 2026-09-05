import json
import inspect
import sqlite3
import threading
import zipfile
from dataclasses import replace
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook

import mentor.datasets as dataset_module
from mentor.analysis import (
    EphemeralQualitativeEvidence,
    QualitativeDisclosureCapability,
    QualitativeEvidenceMetadata,
    TextEvidenceUseGuard,
    qualitative_evidence_audit_metadata,
    read_text_evidence,
    validate_text_evidence_request,
)
from mentor.datasets import (
    AUTO_MAPPING_POLICY_VERSION,
    continue_qualitative_model_transport,
    DatasetImportError,
    MappingEntry,
    MappingSuggestion,
    create_inspected_mapping_draft,
    inspect_local_dataset,
    import_local_dataset,
    ensure_current_auto_mapping,
    mapping_suggestions,
    model_mapping_context,
    safe_auto_mapping,
)
from mentor.datasets import QualitativeTransportError
from mentor.chat_service import ChatService, EvaluationConfig
from mentor.storage import Storage


class _SentinelResponses:
    """In-memory provider boundary used only to inspect immediate transport input."""

    def __init__(self, *, failures: int = 0):
        self.calls: list[dict[str, object]] = []
        self.failures = failures

    def create(self, **request):
        self.calls.append(request)
        if self.failures:
            self.failures -= 1
            raise TimeoutError("temporary provider failure")
        return type("Response", (), {"status": "completed"})()


class _SequenceResponses:
    def __init__(self, *responses):
        self.responses = list(responses)
        self.calls: list[dict[str, object]] = []

    def create(self, **request):
        self.calls.append(request)
        return self.responses.pop(0)


class _ProjectedOutputWrapper:
    """Small SDK-like response item used to exercise the approved projection path."""

    def __init__(self, payload):
        self.payload = payload

    def model_dump(self, *, mode):
        assert mode == "json"
        return self.payload


def _source_response(text: str, annotations: list[dict[str, object]], *, passage: str = "[730.0 --> 756.0] Jacob's original words."):
    return type("Response", (), {
        "status": "completed",
        "id": "response-qualitative",
        "model": "gpt-5.6-sol",
        "usage": {"input_tokens": 1, "output_tokens": 1, "total_tokens": 2},
        "output": [
            {
                "type": "file_search_call",
                "queries": ["Jacob source"],
                "results": [{
                    "file_id": "file_jacob", "filename": "lesson.txt", "text": passage,
                    "attributes": {"year": "2026"},
                }],
            },
            {
                "type": "message", "role": "assistant",
                "content": [{"type": "output_text", "text": text, "annotations": annotations}],
            },
        ],
    })()


def _qualitative_payload(evidence):
    responses = _SentinelResponses()
    continue_qualitative_model_transport(
        client=type("Client", (), {"responses": responses})(),
        request={"model": "fake", "input": []}, call_id="call_qualitative", evidence=evidence,
    )
    return json.loads(responses.calls[0]["input"][-1]["output"])


def test_qualitative_model_transport_keeps_raw_payload_inside_immediate_client_call(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "TRANSPORT-SENTINEL: market looked weak so I entered short."
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,),
        include_approved_notes=True, use_guard=capability,
    )
    responses = _SentinelResponses()

    result = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": responses})(),
        request={"model": "fake", "input": [{"role": "user", "content": "analyse notes"}]},
        call_id="call_notes",
        evidence=evidence,
    )

    assert result.status == "completed"
    assert note not in repr(result)
    assert note not in json.dumps(result.to_persistable_dict())
    sent = json.loads(responses.calls[0]["input"][-1]["output"])
    assert sent["items"][0]["text"][0]["value"] == note
    assert capability.active is False
    assert note not in json.dumps(evidence.to_persistable_metadata().to_dict())


def test_qualitative_transport_returns_safe_continuation_for_same_turn_citation_repair(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    storage.set_vector_store("vs_jacob")
    source = tmp_path / "trades.csv"
    note = "SECRET_NOTE_SENTINEL_REPAIR"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    draft = _source_response("Direct source teaching: Jacob teaches this.", [])
    repaired = _source_response(
        "Direct source teaching: Jacob teaches this.",
        [{"type": "file_citation", "file_id": "file_jacob", "filename": "lesson.txt"}],
    )
    responses = _SequenceResponses(draft, repaired)

    continuation = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": responses})(),
        request={"model": "fake", "input": [], "instructions": "analyse", "tools": []},
        call_id="call_notes", evidence=evidence,
    )

    assert note in responses.calls[0]["input"][-1]["output"]
    assert note not in repr(continuation)
    assert note not in json.dumps(continuation.to_persistable_dict())
    service = ChatService(storage, type("Client", (), {"responses": responses})())
    response, evidence_output, draft_response = service._citation_repaired_response(
        {"model": "fake", "input": [], "instructions": "analyse", "tools": []},
        continuation,
        "What does Jacob teach?",
    )
    thread_id = storage.create_thread("Qualitative")
    answer = service._finalize(
        thread_id,
        {"role": "user", "content": [{"type": "input_text", "text": "What does Jacob teach?"}]},
        response, EvaluationConfig(), "normal", 0.0, evidence_output=evidence_output, draft_response=draft_response,
    )

    assert len(responses.calls) == 2
    assert "Citation repair" in responses.calls[1]["instructions"]
    assert answer.citations[0].file_id == "file_jacob"
    persisted = json.dumps(
        storage.thread_items(thread_id) + storage.replay_items(thread_id) + storage.response_diagnostics(thread_id)
    )
    assert note not in persisted


def test_qualitative_transport_continuation_supports_same_turn_timestamp_repair_without_raw_replay(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    storage.set_vector_store("vs_jacob")
    source = tmp_path / "trades.csv"
    note = "SECRET_NOTE_SENTINEL_TIMESTAMP"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    citation = [{"type": "file_citation", "file_id": "file_jacob", "filename": "lesson.txt"}]
    draft = _source_response(
        "Direct source teaching: Jacob says this at 12:10–12:36.", citation,
        passage="[609.0 --> 616.0] An unrelated passage.",
    )
    repaired = _source_response("Direct source teaching: Jacob says this at 12:10–12:36.", citation)
    responses = _SequenceResponses(draft, repaired)
    request = {
        "model": "fake", "input": [], "instructions": "analyse",
        "tools": [{"type": "file_search", "vector_store_ids": ["vs_jacob"]}],
    }

    continuation = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": responses})(),
        request=request,
        call_id="call_notes", evidence=evidence,
    )
    service = ChatService(storage, type("Client", (), {"responses": responses})())
    response, evidence_output, draft_response = service._citation_repaired_response(
        request,
        continuation,
        "Where exactly does Jacob say this? Give me the timestamp.",
    )

    assert response is repaired
    assert draft_response is continuation
    assert len(evidence_output) == 4
    assert responses.calls[1]["tool_choice"] == {"type": "file_search"}
    assert note not in json.dumps(responses.calls[1])
    assert note not in json.dumps(continuation.to_persistable_dict())


def test_qualitative_transport_no_repair_and_failed_repair_persist_no_raw_payload(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    storage.set_vector_store("vs_jacob")
    source = tmp_path / "trades.csv"
    note = "SECRET_NOTE_SENTINEL_REPAIR_FAILURE"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    citation = [{"type": "file_citation", "file_id": "file_jacob", "filename": "lesson.txt"}]

    no_repair_evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    no_repair_responses = _SequenceResponses(_source_response("Direct source teaching: Jacob teaches this.", citation))
    no_repair = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": no_repair_responses})(),
        request={"model": "fake", "input": [], "instructions": "analyse", "tools": []},
        call_id="call_notes", evidence=no_repair_evidence,
    )
    service = ChatService(storage, type("Client", (), {"responses": no_repair_responses})())
    response, _, draft_response = service._citation_repaired_response(
        {"model": "fake", "input": [], "instructions": "analyse", "tools": []}, no_repair, "What does Jacob teach?"
    )
    assert response is no_repair
    assert draft_response is None
    assert len(no_repair_responses.calls) == 1

    failure_evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    failed_repair = type("Response", (), {"status": "failed", "output": []})()
    failure_responses = _SequenceResponses(_source_response("Direct source teaching: Jacob teaches this.", []), failed_repair)
    continuation = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": failure_responses})(),
        request={"model": "fake", "input": [], "instructions": "analyse", "tools": []},
        call_id="call_notes", evidence=failure_evidence,
    )
    service = ChatService(storage, type("Client", (), {"responses": failure_responses})())
    response, evidence_output, draft_response = service._citation_repaired_response(
        {"model": "fake", "input": [], "instructions": "analyse", "tools": []}, continuation, "What does Jacob teach?"
    )
    thread_id = storage.create_thread("Failed repair")
    service._finalize(
        thread_id, {"role": "user", "content": [{"type": "input_text", "text": "What does Jacob teach?"}]},
        response, EvaluationConfig(), "normal", 0.0, evidence_output=evidence_output, draft_response=draft_response,
    )

    assert response.status == "failed"
    assert note not in json.dumps(failure_responses.calls[1])
    persisted = json.dumps(storage.thread_items(thread_id) + storage.replay_items(thread_id) + storage.response_diagnostics(thread_id))
    assert note not in persisted


def test_qualitative_transport_rejects_returned_raw_tool_artifact_and_releases_capability(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "SECRET_NOTE_SENTINEL_MALFORMED"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=capability,
    )
    malformed = type("Response", (), {
        "status": "completed",
        "output": [{"type": "function_call_output", "output": note}],
    })()

    with pytest.raises(QualitativeTransportError) as error:
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": _SequenceResponses(malformed)})(),
            request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
        )

    assert note not in str(error.value)
    assert capability.active is False


@pytest.mark.parametrize(
    "malformed_item",
    [
        {"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"},
        {"type": "message", "content": {"tool": {"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"}}},
        {"type": "message", "content": {"a": {"b": {"c": {"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"}}}}},
        {"type": "message", "content": [{"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"}]},
        {"type": "message", "content": {"children": [{"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"}]}},
        {"type": "message", "content": ({"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"},)},
        _ProjectedOutputWrapper({"type": "message", "content": {"output": {"type": "function_call_output", "output": "SECRET_NOTE_SENTINEL_NESTED_FCO"}}}),
    ],
    ids=("top-level", "nested-dict", "deep-dict", "list", "list-in-dict", "tuple", "sdk-wrapper"),
)
def test_qualitative_transport_rejects_function_call_output_at_every_projected_depth(tmp_path, malformed_item):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Journal\nSECRET_NOTE_SENTINEL_NESTED_FCO\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=capability,
    )
    malformed = type("Response", (), {"status": "completed", "output": [malformed_item]})()

    with pytest.raises(QualitativeTransportError, match="unsafe tool artifact") as error:
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": _SequenceResponses(malformed)})(),
            request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
        )

    assert "SECRET_NOTE_SENTINEL_NESTED_FCO" not in str(error.value)
    assert capability.active is False
    with pytest.raises(ValueError, match="expired"):
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": _SentinelResponses()})(),
            request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
        )


def test_qualitative_transport_accepts_nested_assistant_and_reasoning_state(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Journal\nSAFE_NESTED_CONTINUATION\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=capability,
    )
    safe = type("Response", (), {"status": "completed", "output": [
        _ProjectedOutputWrapper({"type": "reasoning", "summary": [{"type": "summary_text", "text": "safe"}]}),
        {"type": "message", "content": [{"type": "output_text", "text": "safe", "annotations": []}]},
    ]})()

    result = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": _SequenceResponses(safe)})(),
        request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
    )

    assert [item["type"] for item in result.output] == ["reasoning", "message"]
    assert capability.active is False


def test_qualitative_transport_rejects_ephemeral_evidence_wrapped_in_model_output(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Journal\nSECRET_NOTE_SENTINEL_WRAPPED\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=capability,
    )
    malformed = type("Response", (), {
        "status": "completed",
        "output": [{"type": "message", "content": [{"evidence": evidence}]}],
    })()

    with pytest.raises(QualitativeTransportError, match="unsafe"):
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": _SequenceResponses(malformed)})(),
            request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
        )

    assert capability.active is False


def test_qualitative_transport_retries_only_in_memory_and_releases_on_terminal_failure(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "TRANSPORT-RETRY-SENTINEL"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,),
        include_approved_notes=True, use_guard=capability,
    )
    responses = _SentinelResponses(failures=1)

    result = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": responses})(),
        request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence, max_attempts=2,
    )

    assert result.status == "completed"
    assert len(responses.calls) == 2
    assert all(note in call["input"][-1]["output"] for call in responses.calls)
    assert note not in json.dumps(result.to_persistable_dict())
    assert capability.active is False
    with pytest.raises(ValueError, match="expired"):
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": responses})(), request={"model": "fake", "input": []},
            call_id="call_notes", evidence=evidence,
        )

    fresh_capability = QualitativeDisclosureCapability()
    fresh_evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,),
        include_approved_notes=True, use_guard=fresh_capability,
    )
    failing = _SentinelResponses(failures=1)
    with pytest.raises(QualitativeTransportError) as error:
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": failing})(), request={"model": "fake", "input": []},
            call_id="call_notes", evidence=fresh_evidence,
        )
    assert fresh_capability.active is False
    assert note not in repr(fresh_evidence)
    assert note not in str(error.value)


def test_qualitative_transport_result_has_no_generic_raw_serialization_route(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "TRANSPORT-SERIALIZATION-SENTINEL"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,), include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    result = continue_qualitative_model_transport(
        client=type("Client", (), {"responses": _SentinelResponses()})(),
        request={"model": "fake", "input": []}, call_id="call_notes", evidence=evidence,
    )

    with pytest.raises(TypeError):
        json.dumps(result)
    assert note not in json.dumps(result.to_persistable_dict())


def _dataset(storage: Storage):
    return storage.create_dataset(
        dataset_id="dataset-alpha",
        original_name="trades.csv",
        content_sha256="a" * 64,
        original_extension=".csv",
        byte_size=42,
        source_row_count=3,
        status="ready",
        import_spec={
            "header_row": 0,
            "csv_encoding": "utf-8",
            "csv_delimiter": ",",
            "csv_quoting": '"',
            "parser_version": "pandas-3.0.5",
            "row_order_policy": "source",
            "time_parse_policy": "unambiguous_only",
        },
        columns=[
            {"ordinal": 0, "original_header": "Result_R", "inferred_type": "number", "null_count": 0, "invalid_count": 0},
            {"ordinal": 1, "original_header": "Session", "inferred_type": "string", "null_count": 0, "invalid_count": 0},
        ],
    )


def _importer(storage: Storage, tmp_path: Path):
    return lambda source, **kwargs: import_local_dataset(
        source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "dataset-imported", **kwargs
    )


def _workbook(path: Path, *, sheet_name: str = "Trades", formula: bool = False) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Result_R", "Session"])
    worksheet.append(["=1+1" if formula else 1.5, "London"])
    workbook.save(path)


def _rewrite_zip_member(path: Path, member_name: str, transform) -> None:
    rewritten = path.with_suffix(".rewritten.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(rewritten, "w") as destination:
        for member in source.infolist():
            contents = source.read(member.filename)
            destination.writestr(member, transform(contents) if member.filename == member_name else contents)
    rewritten.replace(path)


def test_local_csv_import_copies_bytes_records_dialect_and_reports_duplicate_rows(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_bytes(b'Result_R;Session\r\n1.5;London\r\n1.5;London\r\n')

    result = _importer(storage, tmp_path)(source)

    assert result.dataset.source_row_count == 2
    assert result.duplicate_row_count == 1
    assert result.dataset.content_sha256 == sha256(source.read_bytes()).hexdigest()
    assert result.original_path.read_bytes() == source.read_bytes()
    assert result.original_path.parent.name == result.dataset.id
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute(
            "SELECT csv_encoding, csv_delimiter, csv_quoting, selected_sheet FROM dataset_import_specs"
        ).fetchone() == ("utf-8", ";", '"', None)
        assert connection.execute("SELECT original_header FROM dataset_columns ORDER BY ordinal").fetchall() == [
            ("Result_R",),
            ("Session",),
        ]


def test_local_xlsx_import_uses_only_the_selected_sheet_and_preserves_original(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.xlsx"
    _workbook(source)
    from openpyxl import load_workbook

    selected = load_workbook(source)
    selected.create_sheet("Ignored")
    selected["Ignored"].append(["Other"])
    selected["Ignored"].append(["not imported"])
    selected.save(source)

    result = _importer(storage, tmp_path)(source, selected_sheet="Trades")

    assert result.dataset.source_row_count == 1
    assert result.duplicate_row_count == 0
    assert result.original_path.read_bytes() == source.read_bytes()
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT selected_sheet FROM dataset_import_specs").fetchone() == ("Trades",)
        assert connection.execute("SELECT original_header FROM dataset_columns ORDER BY ordinal").fetchall() == [
            ("Result_R",),
            ("Session",),
        ]


def test_local_inspection_previews_raw_headers_and_values_before_any_mapping(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Session,Trade Date\n1.5,London,2026-01-02\n,New York,01/02/2026\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    inspection = inspect_local_dataset(storage, dataset.id)

    assert inspection.mapping_version is None
    assert inspection.preview == [
        {"Result_R": "1.5", "Session": "London", "Trade Date": "2026-01-02"},
        {"Result_R": "", "Session": "New York", "Trade Date": "01/02/2026"},
    ]
    assert [(column.original_header, column.value_type, column.blank_count, column.invalid_count) for column in inspection.columns] == [
        ("Result_R", "number", 1, 0),
        ("Session", "categorical", 0, 0),
        ("Trade Date", "datetime", 0, 1),
    ]
    assert inspection.columns[2].unavailable_reason == "ambiguous_date"
    assert {(suggestion.column_ordinal, suggestion.semantic_role, suggestion.unit) for suggestion in mapping_suggestions(inspection)} == {
        (0, "trade_return", "R"),
        (1, "session", None),
        (2, "trade_timestamp", None),
    }


def test_local_inspection_keeps_blank_and_invalid_numeric_cells_in_health(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Note\n1.5,ok\nnot-a-number,bad\n,blank\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    column = inspect_local_dataset(storage, dataset.id).columns[0]

    assert (column.value_type, column.valid_count, column.blank_count, column.invalid_count) == ("number", 1, 1, 1)
    assert column.unavailable_reason == "invalid_values_excluded"


def test_safe_auto_mapping_accepts_exact_r_headers_with_mostly_valid_numeric_cells(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "phase5_mock_backtest.csv"
    result_values = ["2"] * 95 + ["2.75"] + ["-0.75"] * 59 + ["0"] * 2 + ["two R", "?", "error"]
    mfe_values = ["3"] * 157 + ["", "", "error"]
    mae_values = ["-1"] * 158 + ["", "error"]
    outcomes = ["Win"] * 96 + ["Loss"] * 59 + ["BE"] * 5
    source.write_text(
        "Result_R,MFE_R,MAE_R,Outcome\n"
        + "".join(f"{result},{mfe},{mae},{outcome}\n" for result, mfe, mae, outcome in zip(result_values, mfe_values, mae_values, outcomes, strict=True)),
        encoding="utf-8",
    )
    dataset = _importer(storage, tmp_path)(source).dataset

    inspection = inspect_local_dataset(storage, dataset.id)
    auto_mapping = safe_auto_mapping(inspection)

    assert [(column.original_header, column.value_type, column.valid_count, column.blank_count, column.invalid_count) for column in inspection.columns] == [
        ("Result_R", "number", 157, 0, 3),
        ("MFE_R", "number", 157, 2, 1),
        ("MAE_R", "number", 158, 1, 1),
        ("Outcome", "categorical", 160, 0, 0),
    ]
    assert auto_mapping.ambiguities == []
    assert {(entry.semantic_role, entry.unit) for entry in auto_mapping.entries} == {
        ("trade_return", "R"), ("mfe", "R"), ("mae", "R"), ("trade_outcome", None)
    }
    confirmed = storage.confirm_mapping_version(create_inspected_mapping_draft(storage, inspection, auto_mapping.entries).id)
    assert {entry.semantic_role for entry in storage.mapping_entries(confirmed.id)} == {"trade_return", "mfe", "mae", "trade_outcome"}


def test_safe_auto_mapping_allows_bounded_standard_trading_group_labels(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text(
        "Session,Direction,Instrument,Setup,Quarter,SMT,Rule_Adherence,Mistake_Tag\n"
        "London,Long,ES,Reversal,Q1,true,High,None\n"
        "New York,Short,NQ,Continuation,Q2,false,Medium,Early entry\n",
        encoding="utf-8",
    )
    dataset = _importer(storage, tmp_path)(source).dataset

    draft = create_inspected_mapping_draft(
        storage, inspection := inspect_local_dataset(storage, dataset.id), safe_auto_mapping(inspection).entries
    )
    confirmed = storage.confirm_mapping_version(draft.id)
    entries = {entry.analysis_label or entry.semantic_role: entry for entry in storage.mapping_entries(confirmed.id)}

    assert {name for name, entry in entries.items() if entry.aggregate_labels_allowed} == {
        "Session", "Direction", "Instrument", "Setup", "Quarter", "SMT", "Rule adherence", "Mistake tag"
    }
    assert all(entries[name].value_type in {"categorical", "boolean"} for name in entries)


def test_safe_auto_mapping_marks_exact_note_aliases_eligible_for_one_turn_consent(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Outcome,Trade_Notes\n1,Win,waited for confirmation\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)

    mapping = storage.confirm_mapping_version(
        create_inspected_mapping_draft(storage, inspection, safe_auto_mapping(inspection).entries).id
    )
    notes = next(entry for entry in storage.mapping_entries(mapping.id) if entry.analysis_label == "Trade notes")

    assert mapping.auto_mapping_policy_version == 3
    assert notes.mentor_access == "allow_row_values_when_analysing_notes"
    validate_text_evidence_request(
        storage,
        dataset.id,
        mapping.id,
        text_field_ids=(notes.field_id or "",),
        context_field_ids=(),
        filters=(),
        order_by="source",
    )


def test_qualitative_review_reports_the_100_of_156_partial_boundary(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "phase5_mock_backtest.csv"
    source.write_text(
        "Trade_Notes\n" + "\n".join(f"synthetic note {number}" for number in range(156)) + "\n",
        encoding="utf-8",
    )
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    mapping = storage.confirm_mapping_version(
        create_inspected_mapping_draft(storage, inspection, safe_auto_mapping(inspection).entries).id
    )
    note_field_id = storage.mapping_entries(mapping.id)[0].field_id or ""

    evidence = read_text_evidence(
        storage,
        dataset.id,
        mapping.id,
        text_field_ids=(note_field_id,),
        include_approved_notes=True,
        use_guard=QualitativeDisclosureCapability(),
    )
    metadata = evidence.to_persistable_metadata().to_dict()

    assert metadata["usable_text_rows"] == 156
    assert metadata["returned_rows"] == 100
    assert metadata["omitted_rows"] == 56
    assert metadata["complete"] is False


def test_stale_deterministic_auto_mapping_gets_one_safe_successor_and_resets_model_replay(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text(
        "Result_R,Outcome,Session,Setup\n"
        "2,Win,London,Reversal\n"
        "-1,Loss,New York,Continuation\n",
        encoding="utf-8",
    )
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    legacy_entries = [
        replace(entry, analysis_label=None, model_disclosure=False)
        if entry.semantic_role in {"session", "setup"} else entry
        for entry in safe_auto_mapping(inspection).entries
    ]
    legacy = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspection, legacy_entries, auto_mapping_policy_version=1
    ).id)
    thread_id = storage.create_thread("Existing attachment")
    storage.set_thread_dataset_scope(thread_id, dataset.id)
    storage.replace_replay_items(thread_id, [{"type": "message", "role": "assistant", "content": "Old empirical result."}])

    upgraded, changed = ensure_current_auto_mapping(storage, dataset.id)

    assert changed is True
    assert upgraded.id != legacy.id
    assert upgraded.auto_mapping_policy_version == AUTO_MAPPING_POLICY_VERSION
    entries = {entry.semantic_role: entry for entry in storage.mapping_entries(upgraded.id)}
    assert entries["session"].aggregate_labels_allowed is True
    assert entries["setup"].aggregate_labels_allowed is True
    assert storage.mapping_entries(legacy.id)[2].aggregate_labels_allowed is False
    assert storage.replay_items(thread_id) == []
    assert ensure_current_auto_mapping(storage, dataset.id) == (upgraded, False)


def test_policy_two_auto_note_mapping_gets_one_immutable_policy_three_successor(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Outcome,Trade_Notes\n1,Win,synthetic note\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    legacy_entries = [
        replace(entry, mentor_access="aggregates_only") if entry.analysis_label == "Trade notes" else entry
        for entry in safe_auto_mapping(inspection).entries
    ]
    legacy = storage.confirm_mapping_version(
        create_inspected_mapping_draft(
            storage,
            inspection,
            legacy_entries,
            auto_mapping_policy_version=2,
        ).id
    )

    successor, changed = ensure_current_auto_mapping(storage, dataset.id)
    old_notes = next(entry for entry in storage.mapping_entries(legacy.id) if entry.analysis_label == "Trade notes")
    new_notes = next(entry for entry in storage.mapping_entries(successor.id) if entry.analysis_label == "Trade notes")
    repeated, changed_again = ensure_current_auto_mapping(storage, dataset.id)

    assert changed is True
    assert successor.id != legacy.id
    assert successor.auto_mapping_policy_version == 3
    assert old_notes.mentor_access == "aggregates_only"
    assert new_notes.mentor_access == "allow_row_values_when_analysing_notes"
    assert repeated.id == successor.id
    assert changed_again is False


def test_auto_mapping_upgrade_preserves_manual_entries_and_note_access(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text(
        "Result_R,Outcome,Session,Setup,Journal\n"
        "2,Win,London,Reversal,waited for confirmation\n"
        "-1,Loss,New York,Continuation,entered early\n",
        encoding="utf-8",
    )
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    legacy_entries = []
    for entry in safe_auto_mapping(inspection).entries:
        if entry.semantic_role == "session":
            legacy_entries.append(MappingEntry(2, semantic_role="session", source="manual"))
        elif entry.semantic_role == "setup":
            legacy_entries.append(replace(entry, analysis_label=None, model_disclosure=False))
        elif entry.analysis_label == "Journal":
            legacy_entries.append(replace(entry, source="manual", mentor_access="allow_row_values_when_analysing_notes"))
        else:
            legacy_entries.append(entry)
    legacy = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspection, legacy_entries, auto_mapping_policy_version=1
    ).id)

    upgraded, changed = ensure_current_auto_mapping(storage, dataset.id)

    assert changed is True
    assert upgraded.id != legacy.id
    entries = {entry.semantic_role or entry.analysis_label: entry for entry in storage.mapping_entries(upgraded.id)}
    assert entries["session"].source == "manual"
    assert entries["session"].aggregate_labels_allowed is False
    assert entries["setup"].source == "deterministic_auto"
    assert entries["setup"].aggregate_labels_allowed is True
    assert entries["Journal"].source == "manual"
    assert entries["Journal"].mentor_access == "allow_row_values_when_analysing_notes"


def test_current_auto_mapping_policy_is_a_noop_for_newly_attached_dataset(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Outcome,Session\n1,Win,London\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage, inspection, safe_auto_mapping(inspection).entries
    ).id)

    current, changed = ensure_current_auto_mapping(storage, dataset.id)

    assert mapping.auto_mapping_policy_version == AUTO_MAPPING_POLICY_VERSION
    assert (current, changed) == (mapping, False)


def test_safe_auto_mapping_keeps_high_cardinality_generic_label_non_groupable(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text(
        "Quarter\n" + "".join(f"Q{number}\n" for number in range(1, 22)), encoding="utf-8"
    )
    dataset = _importer(storage, tmp_path)(source).dataset

    entry = safe_auto_mapping(inspect_local_dataset(storage, dataset.id)).entries[0]

    assert entry.analysis_label == "Quarter"
    assert entry.model_disclosure is False


def test_confirming_replacement_mapping_resets_only_model_replay(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    initial = storage.confirm_mapping_version(storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    ).id)
    thread_id = storage.create_thread("Replay boundary")
    storage.set_thread_dataset_scope(thread_id, dataset.id)
    storage.append_thread_items(thread_id, [{"type": "message", "role": "assistant", "content": "Old empirical conclusion."}])
    storage.replace_replay_items(thread_id, [{"type": "message", "role": "assistant", "content": "Old empirical conclusion."}])

    replacement = storage.confirm_mapping_version(storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    ).id)

    assert replacement.version == initial.version + 2
    assert storage.replay_items(thread_id) == []
    assert storage.thread_items(thread_id) == [{"type": "message", "role": "assistant", "content": "Old empirical conclusion."}]


def test_safe_auto_mapping_refuses_exact_r_header_with_contradictory_values(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "contradictory.csv"
    source.write_text("Result_R,Notes\ngreat,one\nbad,two\nmaybe,three\nyes,four\nno,five\n1,six\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    inspection = inspect_local_dataset(storage, dataset.id)
    auto_mapping = safe_auto_mapping(inspection)

    assert inspection.columns[0].value_type == "categorical"
    assert not any(entry.semantic_role == "trade_return" for entry in auto_mapping.entries)
    assert auto_mapping.ambiguities == [{"column_ordinal": 0, "header": "Result_R", "role": "trade_return"}]


def test_safe_auto_mapping_requires_ninety_percent_valid_values_for_exact_r_headers(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    accepted = tmp_path / "accepted.csv"
    accepted.write_text("Result_R\n" + "".join(f"{value}\n" for value in [*range(9), "bad"]), encoding="utf-8")
    rejected = tmp_path / "rejected.csv"
    rejected.write_text("Result_R\n" + "".join(f"{value}\n" for value in [*range(8), "bad", "error"]), encoding="utf-8")

    accepted_dataset = _importer(storage, tmp_path)(accepted).dataset
    rejected_storage = Storage(tmp_path / "rejected" / "mentor.sqlite3")
    rejected_storage.initialize()
    rejected_dataset = _importer(rejected_storage, tmp_path / "rejected")(rejected).dataset
    accepted_inspection = inspect_local_dataset(storage, accepted_dataset.id)
    rejected_inspection = inspect_local_dataset(rejected_storage, rejected_dataset.id)

    assert (accepted_inspection.columns[0].value_type, accepted_inspection.columns[0].valid_count, accepted_inspection.columns[0].invalid_count) == ("number", 9, 1)
    assert {(entry.semantic_role, entry.unit) for entry in safe_auto_mapping(accepted_inspection).entries} == {("trade_return", "R")}
    assert rejected_inspection.columns[0].value_type == "categorical"
    assert safe_auto_mapping(rejected_inspection).ambiguities == [{"column_ordinal": 0, "header": "Result_R", "role": "trade_return"}]


def test_small_mixed_notes_column_never_becomes_numeric(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "notes.csv"
    source.write_text("Notes\n1\nword\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    column = inspect_local_dataset(storage, dataset.id).columns[0]

    assert (column.value_type, column.valid_count, column.invalid_count) == ("categorical", 2, 0)


def test_local_inspection_uses_controlled_outcome_normalization(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Outcome\nWin\nloss\nBE\nunknown\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    column = inspect_local_dataset(storage, dataset.id).columns[0]

    assert (column.value_type, column.valid_count, column.invalid_count, column.distinct_count) == ("categorical", 3, 1, 3)
    assert column.unavailable_reason == "invalid_values_excluded"


def test_mapping_draft_needs_confirmation_and_exposes_only_safe_opaque_model_fields(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R,Desk Secret\n1.5,London\n-0.5,New York\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)

    draft = create_inspected_mapping_draft(
        storage,
        inspection,
        [
            MappingEntry(0, semantic_role="trade_return", unit="R", source="alias"),
            MappingEntry(1, analysis_label="Trading session", source="manual"),
        ],
    )

    assert draft.status == "draft"
    assert model_mapping_context(storage, draft.id) == []
    confirmed = storage.confirm_mapping_version(draft.id)
    context = model_mapping_context(storage, confirmed.id)
    assert context == [
        {
            "field_id": context[0]["field_id"],
            "semantic_role": "trade_return",
            "label": None,
            "value_type": "number",
            "unit": "R",
            "health": {"valid_count": 2, "blank_count": 0, "invalid_count": 0, "ambiguous_date_count": 0, "unavailable_reason": None},
            "aggregate_labels_allowed": False,
            "mentor_access": "aggregates_only",
        },
        {
            "field_id": context[1]["field_id"],
            "label": "Trading session",
            "value_type": "categorical",
            "semantic_role": None,
            "unit": None,
            "health": {"valid_count": 2, "blank_count": 0, "invalid_count": 0, "ambiguous_date_count": 0, "unavailable_reason": None},
            "aggregate_labels_allowed": False,
            "mentor_access": "aggregates_only",
        }
    ]
    assert all(item["field_id"].startswith("field_") for item in context)
    assert "Desk Secret" not in json.dumps(context)
    assert "Result_R" not in json.dumps(context)


def test_mapping_text_access_is_default_denied_and_persisted_immutably(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Journal,Session\nsynthetic note,London\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)

    default_draft = create_inspected_mapping_draft(
        storage, inspection, [MappingEntry(0, analysis_label="Journal")]
    )
    approved_draft = create_inspected_mapping_draft(
        storage,
        inspection,
        [
            MappingEntry(
                0,
                analysis_label="Journal",
                mentor_access="allow_row_values_when_analysing_notes",
            )
        ],
    )

    assert storage.mapping_entries(default_draft.id)[0].mentor_access == "aggregates_only"
    approved = storage.confirm_mapping_version(approved_draft.id)
    assert storage.mapping_entries(approved.id)[0].mentor_access == "allow_row_values_when_analysing_notes"


def test_text_evidence_is_ephemeral_and_cannot_enter_analysis_persistence(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "synthetic-private-note-ephemeral"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    journal_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(journal_id,),
        include_approved_notes=True, use_guard=TextEvidenceUseGuard(),
    )
    thread_id = storage.create_thread("Qualitative evidence")

    assert _qualitative_payload(evidence)["items"][0]["text"][0]["value"] == note
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=mapping.id,
            operation="read_text_evidence",
            schema_version="1.0",
            arguments={"dataset_id": dataset.id},
            result=evidence,
        )
    with sqlite3.connect(storage.database_path) as connection:
        assert note not in "\n".join(connection.iterdump())
        for table in (
            "thread_items",
            "thread_replay_items",
            "response_diagnostics",
            "display_turns",
            "analysis_evidence",
            "analysis_tool_outputs",
        ):
            assert note not in repr(connection.execute(f"SELECT * FROM {table}").fetchall())
    assert storage.analysis_evidence(thread_id) == []
    assert storage.analysis_tool_outputs(thread_id) == []


def test_returned_text_evidence_is_rejected_from_replay_and_diagnostics(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "synthetic-ephemeral-replay-note"
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(storage.mapping_entries(mapping.id)[0].field_id or "",),
        include_approved_notes=True, use_guard=TextEvidenceUseGuard(),
    )
    thread_id = storage.create_thread("Ephemeral replay")
    audit = qualitative_evidence_audit_metadata(evidence, include_approved_notes=True)
    assert "items" not in audit
    assert note not in json.dumps(audit)

    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.replace_replay_items(thread_id, [{"type": "function_call_output", "output": evidence}])
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.record_response_diagnostics(thread_id, "response-note", {"tool_output": evidence})
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.append_thread_items(thread_id, [{"type": "function_call_output", "output": {"wrapper": evidence}}])
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.record_display_turn(
            thread_id, user_text="Question", answer_markdown="Answer", citations=[], evidence=[evidence], diagnostics=None,
            response_id=None, status="completed", incomplete_reason=None,
        )
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.record_display_turn(
            thread_id, user_text="Question", answer_markdown="Answer", citations=[], evidence=[], diagnostics={"tool_output": evidence},
            response_id=None, status="completed", incomplete_reason=None,
        )
    storage.replace_replay_items(thread_id, [{"type": "compaction", "status": "completed"}])
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.append_replay_items(thread_id, [{"type": "function_call_output", "output": evidence}])
    storage.record_response_diagnostics(thread_id, "response-safe", {"status": "completed", "qualitative_audit": audit})
    storage.append_replay_items(thread_id, [{"type": "function_call_output", "output": audit}])
    storage.record_display_turn(
        thread_id, user_text="Question", answer_markdown=note, citations=[], evidence=[], diagnostics={"qualitative_audit": audit},
        response_id=None, status="completed", incomplete_reason=None,
    )

    storage.record_qualitative_metadata(thread_id, 1, evidence.to_persistable_metadata())

    with sqlite3.connect(storage.database_path) as connection:
        for table, columns in (
            ("thread_items", "item_json"),
            ("thread_replay_items", "item_json"),
            ("response_diagnostics", "diagnostic_json"),
        ):
            assert note not in "\n".join(row[0] for row in connection.execute(f"SELECT {columns} FROM {table}"))
        display_metadata = connection.execute("SELECT evidence_json, diagnostic_json FROM display_turns").fetchall()
        assert note not in repr(display_metadata)
    assert storage.replay_items(thread_id)[-1]["output"] == audit
    assert storage.response_diagnostics(thread_id) == [{"status": "completed", "qualitative_audit": audit}]
    assert storage.display_turns(thread_id)[0]["answer_markdown"] == note
    assert storage.qualitative_metadata(thread_id) == [audit]


def test_qualitative_evidence_is_typed_ephemeral_and_only_metadata_persists(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    note = "Market looked weak so I entered short."
    source.write_text(f"Journal\n{note}\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    mapping = storage.confirm_mapping_version(create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, analysis_label="Journal", mentor_access="allow_row_values_when_analysing_notes")],
    ).id)
    field_id = storage.mapping_entries(mapping.id)[0].field_id or ""
    capability = QualitativeDisclosureCapability()
    evidence = read_text_evidence(
        storage, dataset.id, mapping.id, text_field_ids=(field_id,),
        include_approved_notes=True, use_guard=capability,
    )
    thread_id = storage.create_thread("Typed evidence")

    assert isinstance(evidence, EphemeralQualitativeEvidence)
    assert note not in repr(evidence)
    with pytest.raises(TypeError):
        json.dumps(evidence)
    metadata = evidence.to_persistable_metadata()
    assert isinstance(metadata, QualitativeEvidenceMetadata)
    assert note not in json.dumps(metadata.to_dict())
    assert _qualitative_payload(evidence)["items"][0]["text"][0]["value"] == note

    for wrapped in (evidence, {"result": evidence}, [evidence], {"diagnostic": evidence}):
        with pytest.raises(ValueError, match="ephemeral qualitative"):
            storage.record_response_diagnostics(thread_id, "response-rejected", {"value": wrapped})
    with pytest.raises(ValueError, match="ephemeral qualitative"):
        storage.append_replay_items(thread_id, [{"type": "function_call_output", "output": evidence}])

    storage.record_response_diagnostics(thread_id, "response-safe", {"qualitative_metadata": metadata.to_dict()})
    storage.append_thread_items(thread_id, [{"role": "user", "content": [{"type": "input_text", "text": note}]}])
    capability.release()
    with pytest.raises(ValueError, match="expired"):
        continue_qualitative_model_transport(
            client=type("Client", (), {"responses": _SentinelResponses()})(),
            request={"model": "fake", "input": []}, call_id="call_qualitative", evidence=evidence,
        )
    with pytest.raises(ValueError, match="expired"):
        read_text_evidence(
            storage, dataset.id, mapping.id, text_field_ids=(field_id,),
            include_approved_notes=True, use_guard=capability,
        )

    assert storage.thread_items(thread_id)[0]["content"][0]["text"] == note
    assert storage.response_diagnostics(thread_id) == [{"qualitative_metadata": metadata.to_dict()}]
    storage.record_qualitative_metadata(thread_id, 1, metadata)
    assert storage.qualitative_metadata(thread_id) == [metadata.to_dict()]
    assert storage.delete_thread(thread_id) is True
    assert storage.qualitative_metadata(thread_id) == []


def test_legacy_mapping_migration_defaults_text_access_to_denied(tmp_path):
    database_path = tmp_path / "legacy.sqlite3"
    contents = "Journal\nlegacy synthetic note\n"
    original = tmp_path / "datasets" / "legacy" / "original.csv"
    original.parent.mkdir(parents=True)
    original.write_text(contents, encoding="utf-8")
    with sqlite3.connect(database_path) as connection:
        connection.executescript(
            """
            CREATE TABLE datasets (
                id TEXT PRIMARY KEY, original_name TEXT NOT NULL, content_sha256 TEXT NOT NULL,
                original_extension TEXT NOT NULL, byte_size INTEGER NOT NULL, source_row_count INTEGER NOT NULL,
                status TEXT NOT NULL, imported_at TEXT NOT NULL
            );
            CREATE TABLE dataset_import_specs (
                id INTEGER PRIMARY KEY, dataset_id TEXT NOT NULL UNIQUE, selected_sheet TEXT,
                header_row INTEGER NOT NULL, csv_encoding TEXT, csv_delimiter TEXT, csv_quoting TEXT,
                parser_version TEXT, row_order_policy TEXT, time_parse_policy TEXT, created_at TEXT NOT NULL
            );
            CREATE TABLE dataset_columns (
                dataset_id TEXT NOT NULL, ordinal INTEGER NOT NULL, original_header TEXT NOT NULL,
                inferred_type TEXT NOT NULL, null_count INTEGER NOT NULL, invalid_count INTEGER NOT NULL,
                PRIMARY KEY(dataset_id, ordinal)
            );
            CREATE TABLE dataset_mapping_versions (
                id INTEGER PRIMARY KEY, dataset_id TEXT NOT NULL, version INTEGER NOT NULL,
                status TEXT NOT NULL, created_at TEXT NOT NULL, confirmed_at TEXT, UNIQUE(dataset_id, version)
            );
            CREATE TABLE dataset_mapping_entries (
                mapping_version_id INTEGER NOT NULL, column_ordinal INTEGER NOT NULL, semantic_role TEXT,
                unit TEXT, analysis_label TEXT, source TEXT NOT NULL, field_id TEXT, value_type TEXT,
                valid_count INTEGER NOT NULL, blank_count INTEGER NOT NULL, invalid_count INTEGER NOT NULL,
                distinct_count INTEGER NOT NULL, max_label_length INTEGER NOT NULL,
                aggregate_labels_allowed INTEGER NOT NULL, unavailable_reason TEXT,
                ambiguous_date_count INTEGER NOT NULL, PRIMARY KEY(mapping_version_id, column_ordinal)
            );
            """
        )
        connection.execute(
            "INSERT INTO datasets VALUES ('legacy', 'legacy.csv', ?, '.csv', ?, 1, 'ready', '2026-08-27T00:00:00Z')",
            (sha256(contents.encode()).hexdigest(), len(contents.encode())),
        )
        connection.execute(
            "INSERT INTO dataset_import_specs VALUES (1, 'legacy', NULL, 0, 'utf-8', ',', '\"', 'pandas-3.0.5', 'source', 'unambiguous_only', '2026-08-27T00:00:00Z')"
        )
        connection.execute("INSERT INTO dataset_columns VALUES ('legacy', 0, 'Journal', 'categorical', 0, 0)")
        connection.execute(
            "INSERT INTO dataset_mapping_versions VALUES (1, 'legacy', 1, 'confirmed', '2026-08-27T00:00:00Z', '2026-08-27T00:00:00Z')"
        )
        connection.execute(
            "INSERT INTO dataset_mapping_entries VALUES (1, 0, NULL, NULL, 'Journal', 'manual', 'field_0123456789ab', "
            "'categorical', 1, 0, 0, 1, 22, 0, NULL, 0)"
        )

    storage = Storage(database_path)
    storage.initialize()
    entry = storage.mapping_entries(1)[0]

    assert entry.mentor_access == "aggregates_only"
    with pytest.raises(ValueError, match="text field is not eligible"):
        read_text_evidence(
            storage, "legacy", 1, text_field_ids=(entry.field_id or "",),
            include_approved_notes=True, use_guard=TextEvidenceUseGuard(),
        )


def test_mapping_edits_and_clears_create_new_health_snapshots_and_reject_unsafe_disclosure(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text(
        "Result_R,Session,Flag\n1,London,true\n-1,New York,false\n0,Asia,true\n", encoding="utf-8"
    )
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)
    draft = create_inspected_mapping_draft(
        storage,
        inspection,
            [
                MappingEntry(0, semantic_role="trade_return", unit="R"),
                MappingEntry(1, semantic_role="session", analysis_label="Session", model_disclosure=True),
                MappingEntry(2, analysis_label="Trade flag", model_disclosure=True),
            ],
    )
    confirmed = storage.confirm_mapping_version(draft.id)

    cleared = create_inspected_mapping_draft(
        storage,
        inspection,
        [MappingEntry(0, analysis_label="Return", source="manual")],
    )
    cleared_confirmed = storage.confirm_mapping_version(cleared.id)

    assert cleared_confirmed.version > confirmed.version
    assert storage.mapping_entries(confirmed.id)[0].semantic_role == "trade_return"
    assert storage.mapping_entries(cleared_confirmed.id)[0].semantic_role is None
    assert storage.mapping_entries(confirmed.id)[1].valid_count == 3
    assert storage.mapping_entries(confirmed.id)[1].value_type == "categorical"
    assert storage.mapping_entries(confirmed.id)[1].aggregate_labels_allowed is True
    with pytest.raises(ValueError, match="unique"):
        create_inspected_mapping_draft(
            storage,
            inspection,
            [MappingEntry(0, semantic_role="trade_return", unit="R"), MappingEntry(1, semantic_role="trade_return", unit="R")],
        )
    categories = tmp_path / "categories.csv"
    categories.write_text("Category\n" + "\n".join(f"group-{index}" for index in range(21)) + "\n", encoding="utf-8")
    category_dataset = import_local_dataset(
        categories, storage, dataset_id_factory=lambda: "dataset-categories"
    ).dataset
    with pytest.raises(ValueError, match="at most 20"):
        create_inspected_mapping_draft(
            storage,
            inspect_local_dataset(storage, category_dataset.id),
            [MappingEntry(0, analysis_label="Category", model_disclosure=True)],
        )


def test_mapping_storage_rejects_forged_or_missing_inspection_snapshot_fields(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)

    with pytest.raises(ValueError, match="inspection snapshot"):
        storage.create_mapping_draft(
            dataset.id, [MappingEntry(0, semantic_role="trade_return", unit="R")]
        )
    with pytest.raises(ValueError, match="inspection snapshot"):
        storage.create_mapping_draft(
            dataset.id,
            [
                MappingEntry(
                    0,
                    semantic_role="trade_return",
                    unit="R",
                    field_id="field_forged",
                    value_type="number",
                    valid_count=99,
                )
            ],
        )
    draft = create_inspected_mapping_draft(
        storage, inspection, [MappingEntry(0, semantic_role="trade_return", unit="R")]
    )
    with sqlite3.connect(storage.database_path) as connection:
        connection.execute(
            "UPDATE dataset_mapping_entries SET field_id = 'field_forged' WHERE mapping_version_id = ?",
            (draft.id,),
        )
    with pytest.raises(ValueError, match="inspection snapshot"):
        storage.confirm_mapping_version(draft.id)


def test_mapping_storage_rejects_spoofed_semantic_role_unit_and_source(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    draft = create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, semantic_role="trade_return", unit="R")],
    )
    entry = storage.mapping_entries(draft.id)[0]

    for spoofed in (
        replace(entry, semantic_role="made_up"),
        replace(entry, unit="made_up"),
        replace(entry, source="made_up"),
    ):
        with pytest.raises(ValueError, match="mapping semantic"):
            storage.create_mapping_draft(dataset.id, [spoofed])
    with sqlite3.connect(storage.database_path) as connection:
        connection.execute(
            "UPDATE dataset_mapping_entries SET semantic_role = 'made_up' WHERE mapping_version_id = ?",
            (draft.id,),
        )
    with pytest.raises(ValueError, match="mapping semantic"):
        storage.confirm_mapping_version(draft.id)


def test_manual_trade_outcome_mapping_uses_controlled_values_not_header_alias(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Status\nWin\nunknown\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset
    inspection = inspect_local_dataset(storage, dataset.id)

    assert (inspection.columns[0].valid_count, inspection.columns[0].invalid_count) == (2, 0)
    draft = create_inspected_mapping_draft(
        storage,
        inspection,
        [MappingEntry(0, semantic_role="trade_outcome")],
    )
    entry = storage.mapping_entries(draft.id)[0]

    assert (entry.valid_count, entry.invalid_count, entry.distinct_count) == (1, 1, 1)
    assert entry.unavailable_reason == "invalid_values_excluded"
    assert storage.mapping_entries(storage.confirm_mapping_version(draft.id).id)[0] == entry


def test_explicit_trade_outcome_mapping_overrides_date_like_header_inference(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Outcome Date\nWin\nunknown\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    draft = create_inspected_mapping_draft(
        storage,
        inspect_local_dataset(storage, dataset.id),
        [MappingEntry(0, semantic_role="trade_outcome")],
    )
    entry = storage.mapping_entries(draft.id)[0]

    assert (entry.value_type, entry.valid_count, entry.invalid_count) == ("categorical", 1, 1)
    assert entry.unavailable_reason == "invalid_values_excluded"


def test_tradedate_alias_keeps_ambiguous_and_impossible_dates_distinct(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("TradeDate\n2026-01-02\n01/02/2026\n2026-99-99\n", encoding="utf-8")
    dataset = _importer(storage, tmp_path)(source).dataset

    inspection = inspect_local_dataset(storage, dataset.id)
    column = inspection.columns[0]

    assert (column.value_type, column.valid_count, column.invalid_count, column.ambiguous_date_count) == (
        "datetime",
        1,
        2,
        1,
    )
    assert column.unavailable_reason == "ambiguous_date"
    assert mapping_suggestions(inspection) == [MappingSuggestion(0, "trade_timestamp", None)]


@pytest.mark.parametrize(
    ("filename", "contents", "message"),
    [
        ("bad.csv", b"Result_R\n\xff\n", "decode"),
        ("bad.csv", b'Result_R\n"unterminated\n', "CSV"),
        ("bad.txt", b"Result_R\n1\n", "CSV or XLSX"),
    ],
)
def test_rejected_csv_or_extension_leaves_no_dataset_or_temporary_files(tmp_path, filename, contents, message):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / filename
    source.write_bytes(contents)

    with pytest.raises(DatasetImportError, match=message):
        _importer(storage, tmp_path)(source)

    assert storage.dataset("dataset-imported") is None
    assert not (tmp_path / "datasets").exists() or not list((tmp_path / "datasets").iterdir())


def test_xlsx_formula_macro_external_link_and_archive_limit_are_rejected_before_import(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    formula = tmp_path / "formula.xlsx"
    _workbook(formula, formula=True)
    external_link = tmp_path / "external-link.xlsx"
    _workbook(external_link)
    with zipfile.ZipFile(external_link, "a") as archive:
        archive.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            "<Relationships><Relationship TargetMode='External'/></Relationships>",
        )
    external_target = tmp_path / "external-target.xlsx"
    _workbook(external_target)
    with zipfile.ZipFile(external_target, "a") as archive:
        archive.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            "<Relationships><Relationship Id='rId1' Target='https://example.invalid/data'/></Relationships>",
        )
    macro = tmp_path / "macro.xlsx"
    _workbook(macro)
    with zipfile.ZipFile(macro, "a") as archive:
        archive.writestr("xl/vbaProject.bin", b"not a macro we will execute")
    oversized = tmp_path / "oversized.xlsx"
    _workbook(oversized)

    with pytest.raises(DatasetImportError, match="formula"):
        _importer(storage, tmp_path)(formula)
    with pytest.raises(DatasetImportError, match="external"):
        _importer(storage, tmp_path)(external_link)
    with pytest.raises(DatasetImportError, match="external"):
        _importer(storage, tmp_path)(external_target)
    with pytest.raises(DatasetImportError, match="macro"):
        _importer(storage, tmp_path)(macro)
    monkeypatch.setattr("mentor.datasets.MAX_XLSX_ARCHIVE_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(DatasetImportError, match="archive"):
        _importer(storage, tmp_path)(oversized)

    assert storage.dataset("dataset-imported") is None
    assert not (tmp_path / "datasets").exists() or not list((tmp_path / "datasets").iterdir())


def test_non_xlsx_signature_and_unsafe_generated_identifier_cannot_create_dataset_paths(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    non_xlsx = tmp_path / "not-a-workbook.xlsx"
    non_xlsx.write_bytes(b"not a zip archive")
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")

    with pytest.raises(DatasetImportError, match="signature"):
        _importer(storage, tmp_path)(non_xlsx)
    with pytest.raises(DatasetImportError, match="identifier"):
        import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "../escape")

    assert not (tmp_path / "escape").exists()
    assert storage.dataset("dataset-imported") is None


def test_import_rolls_back_original_when_metadata_recording_fails(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")

    def fail_metadata(**_kwargs):
        raise RuntimeError("database unavailable")

    monkeypatch.setattr(storage, "create_dataset", fail_metadata)

    with pytest.raises(RuntimeError, match="database unavailable"):
        _importer(storage, tmp_path)(source)

    assert not (tmp_path / "datasets").exists() or not list((tmp_path / "datasets").iterdir())


def test_retry_with_existing_dataset_id_preserves_the_existing_immutable_original(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    first = _importer(storage, tmp_path)(source)

    with pytest.raises(DatasetImportError, match="identifier"):
        _importer(storage, tmp_path)(source)

    assert first.original_path.read_bytes() == source.read_bytes()
    assert storage.dataset(first.dataset.id) == first.dataset


def test_xlsx_formula_and_column_limits_cannot_be_hidden_by_forged_dimension(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    formula = tmp_path / "dimension-formula.xlsx"
    _workbook(formula, formula=True)
    _rewrite_zip_member(
        formula,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(b'<dimension ref="A1:B2"/>', b'<dimension ref="A1:A1"/>'),
    )
    columns = tmp_path / "dimension-columns.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([f"column_{index}" for index in range(101)])
    worksheet.append(list(range(101)))
    workbook.save(columns)
    _rewrite_zip_member(
        columns,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(b'<dimension ref="A1:CY2"/>', b'<dimension ref="A1:A1"/>'),
    )

    with pytest.raises(DatasetImportError, match="formula"):
        _importer(storage, tmp_path)(formula)
    with pytest.raises(DatasetImportError, match="column"):
        _importer(storage, tmp_path)(columns)


def test_xlsx_macro_relationship_and_drive_qualified_member_are_rejected_by_preflight(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    macro_relationship = tmp_path / "macro-relationship.xlsx"
    _workbook(macro_relationship)
    _rewrite_zip_member(
        macro_relationship,
        "xl/_rels/workbook.xml.rels",
        lambda xml: xml.replace(
            b"</Relationships>",
            b"<Relationship Type='http://schemas.microsoft.com/office/2006/relationships/vbaProject' "
            b"Target='vbaProject.bin'/></Relationships>",
        ),
    )
    drive_member = tmp_path / "drive-member.xlsx"
    _workbook(drive_member)
    with zipfile.ZipFile(drive_member, "a") as archive:
        archive.writestr("C:/outside.xml", "<outside/>")

    with pytest.raises(DatasetImportError, match="macro"):
        _importer(storage, tmp_path)(macro_relationship)
    with pytest.raises(DatasetImportError, match="archive"):
        _importer(storage, tmp_path)(drive_member)


def test_keyboard_interrupt_cleans_promoted_bytes_and_metadata_before_reraising(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    create_dataset = storage.create_dataset

    def interrupt_after_recording(**kwargs):
        create_dataset(**kwargs)
        raise KeyboardInterrupt()

    monkeypatch.setattr(storage, "create_dataset", interrupt_after_recording)

    with pytest.raises(KeyboardInterrupt):
        _importer(storage, tmp_path)(source)

    assert storage.dataset("dataset-imported") is None
    assert not (tmp_path / "datasets").exists() or not list((tmp_path / "datasets").iterdir())


def test_import_rejects_caller_controlled_or_unc_dataset_roots(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")

    with pytest.raises(DatasetImportError, match="root"):
        import_local_dataset(
            source, storage, tmp_path / "outside", dataset_id_factory=lambda: "dataset-imported"
        )
    with pytest.raises(DatasetImportError, match="root"):
        import_local_dataset(
            source, storage, Path(r"\\server\share\datasets"), dataset_id_factory=lambda: "dataset-imported"
        )

    assert not (tmp_path / "outside").exists()


def test_interrupt_after_durable_completion_preserves_the_committed_dataset_on_next_start(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    complete_import = storage.complete_dataset_import

    def interrupt_after_completion(dataset_id):
        complete_import(dataset_id)
        raise KeyboardInterrupt()

    monkeypatch.setattr(storage, "complete_dataset_import", interrupt_after_completion)

    with pytest.raises(KeyboardInterrupt):
        _importer(storage, tmp_path)(source)

    original = tmp_path / "datasets" / "dataset-imported" / "original.csv"
    assert original.read_bytes() == source.read_bytes()
    assert storage.dataset("dataset-imported") is not None
    monkeypatch.setattr(storage, "complete_dataset_import", complete_import)
    import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "dataset-next")
    assert original.read_bytes() == source.read_bytes()
    assert storage.dataset("dataset-imported") is not None


def test_next_start_recovers_staging_directory_when_interrupted_during_filesystem_cleanup(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    create_dataset = storage.create_dataset

    def interrupt_after_recording(**kwargs):
        create_dataset(**kwargs)
        raise KeyboardInterrupt()

    interrupted_directory = tmp_path / "datasets" / "dataset-imported"
    real_rmtree = __import__("shutil").rmtree

    def interrupt_cleanup(path, *args, **kwargs):
        if Path(path) == interrupted_directory:
            raise KeyboardInterrupt()
        return real_rmtree(path, *args, **kwargs)

    monkeypatch.setattr(storage, "create_dataset", interrupt_after_recording)
    monkeypatch.setattr("mentor.datasets.shutil.rmtree", interrupt_cleanup)

    with pytest.raises(KeyboardInterrupt):
        _importer(storage, tmp_path)(source)

    assert interrupted_directory.exists()
    assert storage.pending_dataset_import_ids() == ["dataset-imported"]
    monkeypatch.setattr("mentor.datasets.shutil.rmtree", real_rmtree)
    monkeypatch.setattr(storage, "create_dataset", create_dataset)
    import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "dataset-next")
    assert not interrupted_directory.exists()
    assert storage.pending_dataset_import_ids() == []


@pytest.mark.parametrize("target", ["//server/share", r"\\server\share", " https://example.invalid", "\thttps://example.invalid"])
def test_xlsx_rejects_authority_and_whitespace_prefixed_external_relationship_targets(tmp_path, target):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "external-target.xlsx"
    _workbook(source)
    with zipfile.ZipFile(source, "a") as archive:
        archive.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            f"<Relationships><Relationship Id='rId1' Target='{target}'/></Relationships>",
        )

    with pytest.raises(DatasetImportError, match="external|relationship"):
        _importer(storage, tmp_path)(source)


def test_next_start_recovers_rename_then_interrupt_staging_orphan_without_touching_completed_dataset(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    completed = import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "dataset-keep")
    incomplete_directory = tmp_path / "datasets" / "dataset-imported"
    real_replace = Path.replace
    real_rmtree = __import__("shutil").rmtree

    def rename_then_interrupt(path, target):
        result = real_replace(path, target)
        if Path(target) == incomplete_directory:
            raise KeyboardInterrupt()
        return result

    def leave_staging_directory(path, *args, **kwargs):
        if Path(path) == incomplete_directory:
            raise OSError("simulate process shutdown during cleanup")
        return real_rmtree(path, *args, **kwargs)

    monkeypatch.setattr(Path, "replace", rename_then_interrupt)
    monkeypatch.setattr("mentor.datasets.shutil.rmtree", leave_staging_directory)

    with pytest.raises(KeyboardInterrupt):
        _importer(storage, tmp_path)(source)

    assert (incomplete_directory / ".pending-import").exists()
    assert storage.pending_dataset_import_ids() == ["dataset-imported"]
    monkeypatch.setattr(Path, "replace", real_replace)
    monkeypatch.setattr("mentor.datasets.shutil.rmtree", real_rmtree)
    import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: "dataset-next")

    assert not incomplete_directory.exists()
    assert storage.pending_dataset_import_ids() == []
    assert completed.original_path.read_bytes() == source.read_bytes()
    assert storage.dataset(completed.dataset.id) == completed.dataset


def test_xlsx_rejects_duplicate_relationship_id_that_hides_a_formula_target(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "duplicate-relationship.xlsx"
    _workbook(source)
    with zipfile.ZipFile(source) as archive:
        formula_sheet = archive.read("xl/worksheets/sheet1.xml").replace(
            b"<v>1.5</v>", b"<f>1+1</f><v>2</v>"
        )
    _rewrite_zip_member(
        source,
        "xl/_rels/workbook.xml.rels",
        lambda xml: xml.replace(
            b'Target="/xl/worksheets/sheet1.xml"', b'Target="/xl/worksheets/formula.xml"'
        ).replace(
            b"</Relationships>",
            b'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            b'Id="rId1" Target="/xl/worksheets/sheet1.xml"/></Relationships>',
        ),
    )
    with zipfile.ZipFile(source, "a") as archive:
        archive.writestr("xl/worksheets/formula.xml", formula_sheet)

    with pytest.raises(DatasetImportError, match="relationship"):
        _importer(storage, tmp_path)(source)


def test_overlapping_import_cannot_recover_an_active_staging_dataset(tmp_path, monkeypatch):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    source = tmp_path / "trades.csv"
    source.write_text("Result_R\n1.5\n", encoding="utf-8")
    entered_create = threading.Event()
    release_create = threading.Event()
    create_dataset = storage.create_dataset
    recover_pending_imports = dataset_module._recover_pending_dataset_imports
    recovery_seen = threading.Event()
    failures: list[BaseException] = []

    def block_first_create(**kwargs):
        if kwargs["dataset_id"] == "dataset-first":
            entered_create.set()
            assert release_create.wait(10)
        return create_dataset(**kwargs)

    def run_import(dataset_id):
        try:
            import_local_dataset(source, storage, tmp_path / "datasets", dataset_id_factory=lambda: dataset_id)
        except BaseException as error:
            failures.append(error)

    def watch_recovery(root, active_storage):
        if "dataset-first" in active_storage.pending_dataset_import_ids():
            recovery_seen.set()
        return recover_pending_imports(root, active_storage)

    monkeypatch.setattr(storage, "create_dataset", block_first_create)
    monkeypatch.setattr(dataset_module, "_recover_pending_dataset_imports", watch_recovery)
    first = threading.Thread(target=run_import, args=("dataset-first",))
    second = threading.Thread(target=run_import, args=("dataset-second",))
    first.start()
    assert entered_create.wait(10)
    second.start()
    try:
        assert not recovery_seen.wait(0.25)
    finally:
        release_create.set()
    first.join(10)
    second.join(10)

    assert not first.is_alive()
    assert not second.is_alive()
    assert failures == []
    for dataset_id in ("dataset-first", "dataset-second"):
        assert storage.dataset(dataset_id) is not None
        assert (tmp_path / "datasets" / dataset_id / "original.csv").read_bytes() == source.read_bytes()


def _result_envelope(dataset, mapping_version_id: int, operation: str = "summarize_results"):
    return {
        "provenance": "USER_EMPIRICAL_EVIDENCE",
        "dataset_id": dataset.id,
        "dataset_sha256": dataset.content_sha256,
        "mapping_version_id": mapping_version_id,
        "operation": operation,
        "schema_version": "1",
        "filters": [],
        "metric_definitions": {
            "outcome_rate_denominator": "wins + losses + breakevens",
            "win_rate_interval": "Wilson 95% interval",
            "quantile_method": "linear",
            "return_unit": "R",
            "row_order": "source",
        },
        "counts": {"source_rows": 3, "filtered_rows": 3, "valid_rows": 3, "excluded_rows": 0},
        "disposition_counts": {
            "valid_for_analysis": 3,
            "filtered_out": 0,
            "filter_invalid": 0,
            "required_role_blank": 0,
            "required_role_invalid": 0,
        },
        "exclusion_contract": {
            "row_dispositions_exclusive": True,
            "diagnostic_exclusions_exclusive": False,
            "diagnostic_exclusions_may_overlap": True,
        },
        "exclusions": [],
        "metrics": {"valid_rows": 3},
        "limitations": [],
    }


def test_dataset_metadata_is_immutable_and_never_persists_raw_rows(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()

    dataset = _dataset(storage)

    assert storage.dataset(dataset.id) == dataset
    assert dataset.import_spec_id is not None
    with sqlite3.connect(storage.database_path) as connection:
        assert "raw" not in " ".join(
            row[1].lower()
            for table in ("datasets", "dataset_import_specs", "dataset_columns")
            for row in connection.execute(f"PRAGMA table_info({table})")
        )
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute("UPDATE datasets SET original_name = 'changed.csv' WHERE id = ?", (dataset.id,))


def test_mapping_confirmation_copies_an_atomic_immutable_snapshot_and_blocks_drafts_from_analysis(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    draft = storage.create_mapping_draft(
        dataset.id,
        [
            {"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"},
            {"column_ordinal": 1, "semantic_role": "session", "source": "manual"},
        ],
    )

    draft_thread = storage.create_thread("Draft analysis")
    with pytest.raises(ValueError, match="confirmed"):
        storage.record_analysis_evidence(
            thread_id=draft_thread,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=draft.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=_result_envelope(dataset, draft.id),
        )

    confirmed = storage.confirm_mapping_version(draft.id)

    assert draft.status == "draft"
    assert confirmed.status == "confirmed"
    assert confirmed.id != draft.id
    assert storage.mapping_entries(confirmed.id) == storage.mapping_entries(draft.id)
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="provenance|confirmed"):
            connection.execute(
                "INSERT INTO analysis_evidence(thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id, "
                "mapping_version_id, operation, schema_version, arguments_json, result_json) "
                "VALUES (?, 1, ?, ?, ?, ?, 'summarize_results', '1', '{\"dataset_id\":\"dataset-alpha\"}', '{}')",
                (draft_thread, dataset.id, dataset.content_sha256, dataset.import_spec_id, draft.id),
            )
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute("UPDATE dataset_mapping_versions SET status = 'draft' WHERE id = ?", (confirmed.id,))
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute(
                "INSERT INTO dataset_mapping_entries(mapping_version_id, column_ordinal, semantic_role, unit, analysis_label, source) "
                "VALUES (?, 1, 'trade_return', 'R', NULL, 'manual')",
                (confirmed.id,),
            )


def test_thread_deletion_removes_only_its_dataset_scope_evidence_and_tool_outputs(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    deleted_thread = storage.create_thread("Delete me")
    retained_thread = storage.create_thread("Keep me")
    storage.set_thread_dataset_scope(deleted_thread, dataset.id)
    storage.set_thread_dataset_scope(retained_thread, dataset.id)
    deleted_evidence = storage.record_analysis_evidence(
        thread_id=deleted_thread,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=_result_envelope(dataset, confirmed.id),
    )
    retained_evidence = storage.record_analysis_evidence(
        thread_id=retained_thread,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=_result_envelope(dataset, confirmed.id),
    )
    storage.record_analysis_tool_output(deleted_thread, "call-delete", deleted_evidence.id, _result_envelope(dataset, confirmed.id))
    storage.record_analysis_tool_output(retained_thread, "call-keep", retained_evidence.id, _result_envelope(dataset, confirmed.id))

    assert storage.delete_thread(deleted_thread) is True
    assert storage.dataset(dataset.id) == dataset
    assert storage.thread_dataset_scope(deleted_thread) is None
    assert storage.analysis_evidence(deleted_thread) == []
    assert storage.analysis_tool_outputs(deleted_thread) == []
    assert storage.thread_dataset_scope(retained_thread).dataset_id == dataset.id
    assert [evidence.id for evidence in storage.analysis_evidence(retained_thread)] == [retained_evidence.id]
    assert [output["tool_call_id"] for output in storage.analysis_tool_outputs(retained_thread)] == ["call-keep"]


def test_evidence_provenance_and_result_envelopes_are_immutable_metadata_only(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Evidence")
    envelope = _result_envelope(dataset, confirmed.id)
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=envelope,
    )
    storage.record_analysis_tool_output(thread_id, "valid-output", evidence.id, envelope)

    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=2,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result={"rows": [{"Result_R": 1}]},
        )
    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_tool_output(thread_id, "raw-output", evidence.id, {"body": "private upload"})

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="provenance"):
            connection.execute(
                "INSERT INTO analysis_evidence(thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id, "
                "mapping_version_id, operation, schema_version, arguments_json, result_json) "
                "VALUES (?, 2, ?, ?, ?, ?, 'summarize_results', '1', '{\"dataset_id\":\"dataset-alpha\"}', ?)",
                (thread_id, dataset.id, "b" * 64, dataset.import_spec_id, confirmed.id, json.dumps(envelope)),
            )
        for import_spec_id, mapping_version_id in ((999, confirmed.id), (dataset.import_spec_id, 999)):
            with pytest.raises(sqlite3.IntegrityError, match="provenance"):
                connection.execute(
                    "INSERT INTO analysis_evidence(thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id, "
                    "mapping_version_id, operation, schema_version, arguments_json, result_json) "
                    "VALUES (?, 2, ?, ?, ?, ?, 'summarize_results', '1', '{\"dataset_id\":\"dataset-alpha\"}', ?)",
                    (thread_id, dataset.id, dataset.content_sha256, import_spec_id, mapping_version_id, json.dumps(envelope)),
                )
        with pytest.raises(sqlite3.IntegrityError, match="envelope"):
            connection.execute(
                "INSERT INTO analysis_evidence(thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id, "
                "mapping_version_id, operation, schema_version, arguments_json, result_json) "
                "VALUES (?, 2, ?, ?, ?, ?, 'summarize_results', '1', '{\"dataset_id\":\"dataset-alpha\"}', ?)",
                (thread_id, dataset.id, dataset.content_sha256, dataset.import_spec_id, confirmed.id, json.dumps({"rows": []})),
            )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute("UPDATE analysis_evidence SET dataset_sha256 = ? WHERE id = ?", ("b" * 64, evidence.id))
        with pytest.raises(sqlite3.IntegrityError, match="analysis result envelope"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'raw-bypass', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps({"rows": [{"Result_R": 1}]}),
                ),
            )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE analysis_tool_outputs SET output_json = ? WHERE thread_id = ? AND tool_call_id = 'valid-output'",
                (json.dumps({"rows": [{"Result_R": 1}]}), thread_id),
            )


def test_mapping_versions_keep_confirmed_parent_lineage_and_recover_existing_rows(tmp_path):
    database_path = tmp_path / "mentor.sqlite3"
    with sqlite3.connect(database_path) as connection:
        connection.executescript(
            """
            CREATE TABLE datasets (
                id TEXT PRIMARY KEY, original_name TEXT NOT NULL, content_sha256 TEXT NOT NULL,
                original_extension TEXT NOT NULL, byte_size INTEGER NOT NULL, source_row_count INTEGER NOT NULL,
                status TEXT NOT NULL, imported_at TEXT NOT NULL
            );
            CREATE TABLE dataset_import_specs (
                id INTEGER PRIMARY KEY, dataset_id TEXT NOT NULL UNIQUE, selected_sheet TEXT,
                header_row INTEGER NOT NULL, csv_encoding TEXT, csv_delimiter TEXT, csv_quoting TEXT,
                parser_version TEXT, row_order_policy TEXT, time_parse_policy TEXT, created_at TEXT NOT NULL
            );
            CREATE TABLE dataset_columns (
                dataset_id TEXT NOT NULL, ordinal INTEGER NOT NULL, original_header TEXT NOT NULL,
                inferred_type TEXT NOT NULL, null_count INTEGER NOT NULL, invalid_count INTEGER NOT NULL,
                PRIMARY KEY(dataset_id, ordinal)
            );
            CREATE TABLE dataset_mapping_versions (
                id INTEGER PRIMARY KEY, dataset_id TEXT NOT NULL, version INTEGER NOT NULL,
                status TEXT NOT NULL, created_at TEXT NOT NULL, confirmed_at TEXT,
                UNIQUE(dataset_id, version)
            );
            CREATE TABLE analysis_tool_outputs (
                thread_id INTEGER NOT NULL, tool_call_id TEXT NOT NULL, evidence_id INTEGER NOT NULL,
                output_json TEXT NOT NULL, PRIMARY KEY(thread_id, tool_call_id)
            );
            INSERT INTO datasets VALUES ('legacy', 'legacy.csv', 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
                '.csv', 1, 1, 'ready', '2026-08-27T00:00:00Z');
            INSERT INTO dataset_import_specs VALUES (1, 'legacy', NULL, 0, 'utf-8', ',', '"', 'pandas-3.0.5',
                'source', 'unambiguous_only', '2026-08-27T00:00:00Z');
            INSERT INTO dataset_columns VALUES ('legacy', 0, 'Result_R', 'number', 0, 0);
            INSERT INTO dataset_mapping_versions VALUES (1, 'legacy', 1, 'confirmed', '2026-08-27T00:00:00Z', '2026-08-27T00:00:00Z');
            """
        )

    storage = Storage(database_path)
    storage.initialize()
    with sqlite3.connect(database_path) as connection:
        assert "parent_mapping_version_id" in {
            row[1] for row in connection.execute("PRAGMA table_info(dataset_mapping_versions)")
        }
        assert "arguments_json" in {row[1] for row in connection.execute("PRAGMA table_info(analysis_tool_outputs)")}
        assert connection.execute(
            "SELECT parent_mapping_version_id FROM dataset_mapping_versions WHERE id = 1"
        ).fetchone() == (None,)

    fresh_storage = Storage(tmp_path / "fresh.sqlite3")
    fresh_storage.initialize()
    dataset = _dataset(fresh_storage)
    draft = fresh_storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )
    confirmed = fresh_storage.confirm_mapping_version(draft.id)

    assert confirmed.parent_mapping_version_id == draft.id
    assert fresh_storage.mapping_version(confirmed.id) == confirmed
    with sqlite3.connect(fresh_storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE dataset_mapping_versions SET parent_mapping_version_id = NULL WHERE id = ?", (confirmed.id,)
            )


def test_mapping_drafts_reject_columns_that_are_not_metadata_for_the_dataset(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)

    with pytest.raises(ValueError, match="existing dataset column"):
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 99, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        )

    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM dataset_mapping_versions").fetchone() == (0,)
    valid_draft = storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="existing dataset column"):
            connection.execute(
                "INSERT INTO dataset_mapping_entries(mapping_version_id, column_ordinal, semantic_role, unit, analysis_label, source) "
                "VALUES (?, 99, 'session', NULL, NULL, 'manual')",
                (valid_draft.id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="existing dataset column"):
            connection.execute(
                "UPDATE dataset_mapping_entries SET column_ordinal = 99 WHERE mapping_version_id = ? AND column_ordinal = 0",
                (valid_draft.id,),
            )


def test_analysis_tool_arguments_are_bounded_metadata_and_not_raw_payloads(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Tool arguments")
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=_result_envelope(dataset, confirmed.id),
    )

    assert "arguments" in inspect.signature(storage.record_analysis_tool_output).parameters
    with pytest.raises(ValueError, match="analysis tool arguments"):
        storage.record_analysis_tool_output(
            thread_id,
            "raw-arguments",
            evidence.id,
            _result_envelope(dataset, confirmed.id),
            arguments={"body": "private upload", "rows": [{"Result_R": 1}]},
        )
    with sqlite3.connect(storage.database_path) as connection:
        assert "arguments_json" in {row[1] for row in connection.execute("PRAGMA table_info(analysis_tool_outputs)")}
        with pytest.raises(sqlite3.IntegrityError, match="arguments"):
            connection.execute(
                "INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json) "
                "VALUES (?, 'raw-arguments-bypass', ?, ?, ?)",
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"body": "private upload", "rows": []}),
                    json.dumps(_result_envelope(dataset, confirmed.id)),
                ),
            )


def test_confirmed_mapping_snapshots_cannot_be_promoted_or_changed_with_direct_sql(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    draft = storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )
    confirmed = storage.confirm_mapping_version(draft.id)
    other_draft = storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 1, "semantic_role": "session", "source": "manual"}]
    )

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE dataset_mapping_versions SET status = 'confirmed', confirmed_at = CURRENT_TIMESTAMP WHERE id = ?",
                (other_draft.id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute("UPDATE dataset_mapping_versions SET status = 'draft' WHERE id = ?", (confirmed.id,))
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE dataset_mapping_versions SET parent_mapping_version_id = ? WHERE id = ?",
                (draft.id, other_draft.id),
            )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute("UPDATE dataset_mapping_versions SET version = 99 WHERE id = ?", (confirmed.id,))
        with pytest.raises(sqlite3.IntegrityError, match="confirmed"):
            connection.execute(
                "UPDATE dataset_mapping_entries SET mapping_version_id = ? WHERE mapping_version_id = ?",
                (confirmed.id, other_draft.id),
            )
        with pytest.raises(sqlite3.IntegrityError, match="confirmed"):
            connection.execute(
                "UPDATE dataset_mapping_entries SET semantic_role = 'session' WHERE mapping_version_id = ?",
                (confirmed.id,),
            )


def test_dataset_columns_cannot_be_deleted_while_a_mapping_references_them(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="referenced"):
            connection.execute("DELETE FROM dataset_columns WHERE dataset_id = ? AND ordinal = 0", (dataset.id,))


def test_result_envelopes_reject_raw_values_hidden_in_limitations_or_metrics(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Envelope leakage")
    envelope = _result_envelope(dataset, confirmed.id)
    leaky_limitations = {**envelope, "limitations": ["trade row: Result_R=2.5"]}
    leaky_metrics = {**envelope, "metrics": {"mean_return": "Result_R=2.5"}}
    arbitrary_metric = {**envelope, "metrics": {"raw_result_r_2_5": 2.5}}
    leaky_filter = {
        **envelope,
        "filters": [{"field_id": "Result_R", "operator": "eq", "value_sha256": "a" * 64}],
    }
    leaky_definition = {
        **envelope,
        "metric_definitions": {**envelope["metric_definitions"], "return_unit": "Result_R=2.5"},
    }

    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=leaky_limitations,
        )
    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=leaky_metrics,
        )
    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=arbitrary_metric,
        )
    for leaky_envelope in (leaky_filter, leaky_definition):
        with pytest.raises(ValueError, match="analysis result envelope"):
            storage.record_analysis_evidence(
                thread_id=thread_id,
                origin_turn_number=1,
                dataset_id=dataset.id,
                mapping_version_id=confirmed.id,
                operation="summarize_results",
                schema_version="1",
                arguments={"dataset_id": dataset.id},
                result=leaky_envelope,
            )
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=envelope,
    )
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="envelope"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'leaky', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps(leaky_limitations),
                ),
            )
        with pytest.raises(sqlite3.IntegrityError, match="envelope"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'leaky-metric', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps(leaky_metrics),
                ),
            )
        with pytest.raises(sqlite3.IntegrityError, match="details|envelope"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'leaky-filter', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps(leaky_filter),
                ),
            )


def test_analysis_envelopes_bind_the_confirmed_mapping_return_unit_in_python_and_sqlite(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "percentage", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Unit guard")
    envelope = _result_envelope(dataset, confirmed.id)
    percentage_envelope = {
        **envelope,
        "metric_definitions": {**envelope["metric_definitions"], "return_unit": "percentage"},
    }

    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=envelope,
        )
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=percentage_envelope,
    )
    forged_r_output = {
        **percentage_envelope,
        "metric_definitions": {**percentage_envelope["metric_definitions"], "return_unit": "R"},
        "metrics": {"cumulative_return": 1.0},
    }
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="details|envelope"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'forged-r-unit', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps(forged_r_output),
                ),
            )


def test_evidence_arguments_are_bounded_metadata_and_not_raw_payloads(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Evidence arguments")
    envelope = _result_envelope(dataset, confirmed.id)

    with pytest.raises(ValueError, match="analysis evidence arguments"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"headers": ["Result_R"], "original_filename": "trades.csv", "rows": [[2.5]]},
            result=envelope,
        )
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="evidence arguments"):
            connection.execute(
                """
                INSERT INTO analysis_evidence(
                    thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id,
                    mapping_version_id, operation, schema_version, arguments_json, result_json
                ) VALUES (?, 1, ?, ?, ?, ?, 'summarize_results', '1', ?, ?)
                """,
                (
                    thread_id,
                    dataset.id,
                    dataset.content_sha256,
                    dataset.import_spec_id,
                    confirmed.id,
                    json.dumps({"body": "private data", "rows": [{"Result_R": 2.5}]}),
                    json.dumps(envelope),
                ),
            )


def test_confirmed_mapping_snapshot_must_exactly_match_its_parent_entries(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    draft = storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )
    confirmed = storage.confirm_mapping_version(draft.id)

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="draft"):
            connection.execute(
                """
                INSERT INTO dataset_mapping_versions(dataset_id, version, status, parent_mapping_version_id)
                VALUES (?, 3, 'confirmed', ?)
                """,
                (dataset.id, draft.id),
            )
        connection.execute(
            """
            INSERT INTO dataset_mapping_versions(dataset_id, version, status, parent_mapping_version_id)
            VALUES (?, 4, 'draft', ?)
            """,
            (dataset.id, draft.id),
        )
        forged_id = connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        connection.execute(
            """
            INSERT INTO dataset_mapping_entries(mapping_version_id, column_ordinal, semantic_role, unit, analysis_label, source)
            VALUES (?, 1, 'session', NULL, NULL, 'manual')
            """,
            (forged_id,),
        )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            connection.execute(
                "UPDATE dataset_mapping_versions SET status = 'confirmed', confirmed_at = CURRENT_TIMESTAMP WHERE id = ?",
                (forged_id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="confirmed"):
            connection.execute(
                "INSERT INTO dataset_mapping_entries(mapping_version_id, column_ordinal, semantic_role, unit, analysis_label, source) "
                "VALUES (?, 1, 'session', NULL, NULL, 'manual')",
                (confirmed.id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="confirmed"):
            connection.execute("DELETE FROM dataset_mapping_entries WHERE mapping_version_id = ?", (confirmed.id,))


def test_migration_backfills_preexisting_tool_outputs_before_immutability_triggers(tmp_path):
    database_path = tmp_path / "legacy-tool-output.sqlite3"
    with sqlite3.connect(database_path) as connection:
        connection.executescript(
            """
            CREATE TABLE threads (id INTEGER PRIMARY KEY, title TEXT NOT NULL);
            CREATE TABLE analysis_evidence (
                id INTEGER PRIMARY KEY, thread_id INTEGER NOT NULL, origin_turn_number INTEGER NOT NULL,
                display_turn_number INTEGER, dataset_id TEXT NOT NULL, dataset_sha256 TEXT NOT NULL,
                import_spec_id INTEGER NOT NULL, mapping_version_id INTEGER NOT NULL, operation TEXT NOT NULL,
                schema_version TEXT NOT NULL, arguments_json TEXT NOT NULL, result_json TEXT NOT NULL
            );
            CREATE TABLE analysis_tool_outputs (
                thread_id INTEGER NOT NULL, tool_call_id TEXT NOT NULL, evidence_id INTEGER NOT NULL,
                output_json TEXT NOT NULL, PRIMARY KEY(thread_id, tool_call_id)
            );
            INSERT INTO threads VALUES (1, 'legacy');
            INSERT INTO analysis_evidence VALUES (
                1, 1, 1, NULL, 'legacy-dataset',
                'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
                1, 1, 'summarize_results', '1', '{"dataset_id":"legacy-dataset"}', '{}'
            );
            INSERT INTO analysis_tool_outputs VALUES (1, 'legacy-call', 1, '{}');
            """
        )

    Storage(database_path).initialize()

    with sqlite3.connect(database_path) as connection:
        assert connection.execute(
            "SELECT arguments_json FROM analysis_tool_outputs WHERE tool_call_id = 'legacy-call'"
        ).fetchone() == ('{"dataset_id":"legacy-dataset","mapping_version_id":1,"operation":"summarize_results"}',)


def test_analysis_identifiers_and_metrics_are_safe_finite_and_unique(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Identifier guards")
    envelope = _result_envelope(dataset, confirmed.id)

    with pytest.raises(ValueError, match="identifier"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="trades.csv",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result=_result_envelope(dataset, confirmed.id, operation="trades.csv"),
        )
    with pytest.raises(ValueError, match="identifier"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="source-file.csv",
            arguments={"dataset_id": dataset.id},
            result={**envelope, "schema_version": "source-file.csv"},
        )
    with pytest.raises(ValueError, match="analysis result envelope"):
        storage.record_analysis_evidence(
            thread_id=thread_id,
            origin_turn_number=1,
            dataset_id=dataset.id,
            mapping_version_id=confirmed.id,
            operation="summarize_results",
            schema_version="1",
            arguments={"dataset_id": dataset.id},
            result={**envelope, "metrics": {"mean_return": float("inf")}},
        )
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=envelope,
    )
    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="identifier"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'unsafe-identifier', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps({**envelope, "operation": "raw-file.csv"}),
                ),
            )
        with pytest.raises(sqlite3.IntegrityError, match="identifier"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'unsafe-version', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps({**envelope, "schema_version": "raw-file.csv"}),
                ),
            )
        duplicate_metrics = json.dumps(envelope).replace(
            '"metrics": {"valid_rows": 3}', '"metrics": {"valid_rows": 3, "valid_rows": 1}'
        )
        with pytest.raises(sqlite3.IntegrityError, match="metrics"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'duplicate-metric', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    duplicate_metrics,
                ),
            )
        with pytest.raises(sqlite3.IntegrityError, match="metrics"):
            connection.execute(
                """
                INSERT INTO analysis_evidence(
                    thread_id, origin_turn_number, dataset_id, dataset_sha256, import_spec_id,
                    mapping_version_id, operation, schema_version, arguments_json, result_json
                ) VALUES (?, 2, ?, ?, ?, ?, 'summarize_results', '1', ?, ?)
                """,
                (
                    thread_id,
                    dataset.id,
                    dataset.content_sha256,
                    dataset.import_spec_id,
                    confirmed.id,
                    json.dumps({"dataset_id": dataset.id}),
                    duplicate_metrics,
                ),
            )
        infinite_metrics = json.dumps(envelope).replace(
            '"metrics": {"valid_rows": 3}', '"metrics": {"mean_return": 1e999}'
        )
        with pytest.raises(sqlite3.IntegrityError, match="metrics"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'infinite-metric', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    infinite_metrics,
                ),
            )


def test_parent_drafts_referenced_by_confirmed_snapshots_are_immutable(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    parent = storage.create_mapping_draft(
        dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
    )
    storage.confirm_mapping_version(parent.id)

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="parent"):
            connection.execute(
                "UPDATE dataset_mapping_entries SET semantic_role = 'session' WHERE mapping_version_id = ?",
                (parent.id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="parent"):
            connection.execute(
                "INSERT INTO dataset_mapping_entries(mapping_version_id, column_ordinal, semantic_role, unit, analysis_label, source) "
                "VALUES (?, 1, 'session', NULL, NULL, 'manual')",
                (parent.id,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="parent"):
            connection.execute("DELETE FROM dataset_mapping_entries WHERE mapping_version_id = ?", (parent.id,))
        with pytest.raises(sqlite3.IntegrityError, match="parent"):
            connection.execute("DELETE FROM dataset_mapping_versions WHERE id = ?", (parent.id,))


def test_dataset_hash_and_tool_call_ids_are_safe_opaque_metadata(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()

    with pytest.raises(ValueError, match="dataset identifier"):
        storage.create_dataset(
            dataset_id="..\\trades.csv",
            original_name="trades.csv",
            content_sha256="a" * 64,
            original_extension=".csv",
            byte_size=1,
            source_row_count=1,
            status="ready",
            import_spec={"header_row": 0},
            columns=[],
        )
    with pytest.raises(ValueError, match="content hash"):
        storage.create_dataset(
            dataset_id="dataset-beta",
            original_name="trades.csv",
            content_sha256="A" * 64,
            original_extension=".csv",
            byte_size=1,
            source_row_count=1,
            status="ready",
            import_spec={"header_row": 0},
            columns=[],
        )

    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Opaque IDs")
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=_result_envelope(dataset, confirmed.id),
    )
    with pytest.raises(ValueError, match="tool call identifier"):
        storage.record_analysis_tool_output(
            thread_id,
            "call/raw-export.csv",
            evidence.id,
            _result_envelope(dataset, confirmed.id),
        )
    storage.record_analysis_tool_output(thread_id, "fc_safe-123", evidence.id, _result_envelope(dataset, confirmed.id))

    with sqlite3.connect(storage.database_path) as connection:
        with pytest.raises(sqlite3.IntegrityError, match="dataset identifier"):
            connection.execute(
                """
                INSERT INTO datasets(id, original_name, content_sha256, original_extension, byte_size, source_row_count, status)
                VALUES ('raw/file.csv', 'trades.csv', ?, '.csv', 1, 1, 'ready')
                """,
                ("b" * 64,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="content hash"):
            connection.execute(
                """
                INSERT INTO datasets(id, original_name, content_sha256, original_extension, byte_size, source_row_count, status)
                VALUES ('dataset-gamma', 'trades.csv', ?, '.csv', 1, 1, 'ready')
                """,
                ("B" * 64,),
            )
        with pytest.raises(sqlite3.IntegrityError, match="tool call identifier"):
            connection.execute(
                """
                INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                VALUES (?, 'call/raw-export.csv', ?, ?, ?)
                """,
                (
                    thread_id,
                    evidence.id,
                    json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                    json.dumps(_result_envelope(dataset, confirmed.id)),
                ),
            )


def test_dataset_and_tool_identifiers_reject_null_and_blob_sql_bypasses(tmp_path):
    storage = Storage(tmp_path / "mentor.sqlite3")
    storage.initialize()
    dataset = _dataset(storage)
    confirmed = storage.confirm_mapping_version(
        storage.create_mapping_draft(
            dataset.id, [{"column_ordinal": 0, "semantic_role": "trade_return", "unit": "R", "source": "manual"}]
        ).id
    )
    thread_id = storage.create_thread("Type guards")
    evidence = storage.record_analysis_evidence(
        thread_id=thread_id,
        origin_turn_number=1,
        dataset_id=dataset.id,
        mapping_version_id=confirmed.id,
        operation="summarize_results",
        schema_version="1",
        arguments={"dataset_id": dataset.id},
        result=_result_envelope(dataset, confirmed.id),
    )
    with pytest.raises(ValueError, match="dataset identifier"):
        storage.create_dataset(
            dataset_id=None,  # type: ignore[arg-type]
            original_name="trades.csv",
            content_sha256="a" * 64,
            original_extension=".csv",
            byte_size=1,
            source_row_count=1,
            status="ready",
            import_spec={"header_row": 0},
            columns=[],
        )
    with pytest.raises(ValueError, match="content hash"):
        storage.create_dataset(
            dataset_id="dataset-bytes",
            original_name="trades.csv",
            content_sha256=b"a" * 64,  # type: ignore[arg-type]
            original_extension=".csv",
            byte_size=1,
            source_row_count=1,
            status="ready",
            import_spec={"header_row": 0},
            columns=[],
        )
    with pytest.raises(ValueError, match="tool call identifier"):
        storage.record_analysis_tool_output(
            thread_id,
            b"call_safe",  # type: ignore[arg-type]
            evidence.id,
            _result_envelope(dataset, confirmed.id),
        )

    with sqlite3.connect(storage.database_path) as connection:
        for dataset_id, content_sha256 in (
            (None, "b" * 64),
            (b"dataset-blob", "b" * 64),
            ("dataset-null-hash", None),
            ("dataset-blob-hash", b"b" * 64),
        ):
            with pytest.raises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO datasets(id, original_name, content_sha256, original_extension, byte_size, source_row_count, status)
                    VALUES (?, 'trades.csv', ?, '.csv', 1, 1, 'ready')
                    """,
                    (dataset_id, content_sha256),
                )
        for tool_call_id in (None, b"call_blob"):
            with pytest.raises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO analysis_tool_outputs(thread_id, tool_call_id, evidence_id, arguments_json, output_json)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        thread_id,
                        tool_call_id,
                        evidence.id,
                        json.dumps({"dataset_id": dataset.id, "mapping_version_id": confirmed.id, "operation": "summarize_results"}),
                        json.dumps(_result_envelope(dataset, confirmed.id)),
                    ),
                )
