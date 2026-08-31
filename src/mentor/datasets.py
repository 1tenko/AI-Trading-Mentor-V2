"""Immutable metadata models and local-only import for backtest datasets."""

import csv
import hashlib
import io
import json
import math
import os
import re
import shutil
import tempfile
import threading
import time
import uuid
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import TYPE_CHECKING, Any, Callable, Iterable, Iterator, Mapping
from xml.etree import ElementTree

from openpyxl import load_workbook

if TYPE_CHECKING:
    from mentor.storage import Storage


MAX_DATASET_BYTES = 50 * 1024 * 1024
MAX_DATASET_ROWS = 250_000
MAX_DATASET_COLUMNS = 100
MAX_DATASET_CELLS = 2_000_000
MAX_XLSX_ARCHIVE_MEMBERS = 1_000
MAX_XLSX_ARCHIVE_COMPRESSED_BYTES = MAX_DATASET_BYTES
MAX_XLSX_ARCHIVE_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
MAX_INSPECTION_PREVIEW_ROWS = 20
MAX_INSPECTION_CELL_CHARS = 200
MAX_MODEL_GROUP_LABELS = 20
MAX_MODEL_GROUP_LABEL_CHARS = 80
MENTOR_ACCESS_POLICIES = frozenset({"aggregates_only", "allow_row_values_when_analysing_notes"})


class QualitativeDisclosureCapability:
    """Server-owned, one-turn permission for local qualitative disclosure."""

    __slots__ = ("_used", "_active", "_transported")

    def __init__(self) -> None:
        self._used = False
        self._active = True
        self._transported = False

    def consume(self) -> None:
        if not self._active:
            raise ValueError("qualitative disclosure capability is expired")
        if self._used:
            raise ValueError("qualitative evidence is limited to one call per turn")
        self._used = True

    def consume_transport(self) -> None:
        if not self._active:
            raise ValueError("qualitative disclosure capability is expired")
        if not self._used or self._transported:
            raise ValueError("qualitative disclosure transport is unavailable")
        self._transported = True

    def release(self) -> None:
        self._active = False

    @property
    def active(self) -> bool:
        return self._active


@dataclass(frozen=True, slots=True)
class QualitativeEvidenceMetadata:
    """The only persistable projection of qualitative source data."""

    _payload: Mapping[str, object] = field(repr=False)

    def to_dict(self) -> dict[str, object]:
        return dict(self._payload)


class EphemeralQualitativeEvidence:
    """In-memory raw qualitative data; deliberately not serializable or persistable."""

    __slots__ = ("_metadata", "_items", "_capability")

    def __init__(
        self,
        *,
        metadata: QualitativeEvidenceMetadata,
        items: list[dict[str, object]],
        capability: QualitativeDisclosureCapability,
    ) -> None:
        self._metadata = metadata
        self._items = items
        self._capability = capability

    def __repr__(self) -> str:
        metadata = self._metadata.to_dict()
        return (
            "<EphemeralQualitativeEvidence "
            f"rows={metadata['returned_rows']} chars={metadata['characters_returned']} "
            f"complete={metadata['complete']}>"
        )

    __str__ = __repr__

    def to_persistable_metadata(self) -> QualitativeEvidenceMetadata:
        return self._metadata

    def _release(self) -> None:
        self._items.clear()
        self._capability.release()


@dataclass(frozen=True, slots=True)
class QualitativeTransportResult:
    """Safe result of one qualitative disclosure; raw text never leaves transport."""

    status: str
    metadata: QualitativeEvidenceMetadata

    def to_persistable_dict(self) -> dict[str, object]:
        return {"status": self.status, "qualitative_metadata": self.metadata.to_dict()}


class QualitativeTransportError(RuntimeError):
    """Safe terminal failure for the raw qualitative model-disclosure boundary."""


def continue_qualitative_model_transport(
    *,
    client: Any,
    request: Mapping[str, object],
    call_id: str,
    evidence: EphemeralQualitativeEvidence,
    max_attempts: int = 1,
) -> QualitativeTransportResult:
    """Send one ephemeral qualitative tool output directly to a Responses continuation.

    The raw payload is built and consumed only in this lexical scope.  No raw
    function output is returned to normal application code.
    """
    if not isinstance(evidence, EphemeralQualitativeEvidence):
        raise ValueError("qualitative transport requires ephemeral evidence")
    if not isinstance(call_id, str) or not call_id:
        raise ValueError("qualitative transport call id is required")
    if not isinstance(request, Mapping) or not isinstance(request.get("input"), list):
        raise ValueError("qualitative transport requires a Responses input list")
    if type(max_attempts) is not int or not 1 <= max_attempts <= 2:
        raise ValueError("qualitative transport attempts must be one or two")

    evidence._capability.consume_transport()
    try:
        raw_output = json.dumps(evidence._metadata.to_dict() | {"items": evidence._items}, separators=(",", ":"))
        continuation_request = dict(request)
        continuation_request["input"] = [
            *request["input"],
            {"type": "function_call_output", "call_id": call_id, "output": raw_output},
        ]
        for attempt in range(max_attempts):
            try:
                response = client.responses.create(**continuation_request)
                status = str(getattr(response, "status", "completed"))
                return QualitativeTransportResult(status=status, metadata=evidence.to_persistable_metadata())
            except TimeoutError:
                if attempt + 1 == max_attempts:
                    raise QualitativeTransportError("qualitative model transport timed out") from None
            except Exception:
                raise QualitativeTransportError("qualitative model transport failed") from None
        raise AssertionError("unreachable")
    finally:
        evidence._release()
_DATASET_ID_PATTERN = re.compile(r"[A-Za-z0-9][A-Za-z0-9_-]{0,79}")
_CELL_REFERENCE_PATTERN = re.compile(r"([A-Z]+)([1-9][0-9]*)$")
_IMPORT_LEASE_LOCK = threading.RLock()


@dataclass(frozen=True)
class Dataset:
    id: str
    original_name: str
    content_sha256: str
    original_extension: str
    byte_size: int
    source_row_count: int
    status: str
    import_spec_id: int


@dataclass(frozen=True)
class DatasetImportSpec:
    header_row: int
    csv_encoding: str | None = None
    csv_delimiter: str | None = None
    csv_quoting: str | None = None
    parser_version: str | None = None
    row_order_policy: str | None = None
    time_parse_policy: str | None = None
    selected_sheet: str | None = None


@dataclass(frozen=True)
class DatasetColumn:
    ordinal: int
    original_header: str
    inferred_type: str
    null_count: int
    invalid_count: int


@dataclass(frozen=True)
class DatasetMappingVersion:
    id: int
    dataset_id: str
    version: int
    status: str
    parent_mapping_version_id: int | None


@dataclass(frozen=True)
class MappingEntry:
    column_ordinal: int
    semantic_role: str | None = None
    unit: str | None = None
    analysis_label: str | None = None
    source: str = "manual"
    field_id: str | None = None
    value_type: str | None = None
    valid_count: int = 0
    blank_count: int = 0
    invalid_count: int = 0
    distinct_count: int = 0
    max_label_length: int = 0
    aggregate_labels_allowed: bool = False
    mentor_access: str = "aggregates_only"
    unavailable_reason: str | None = None
    model_disclosure: bool = False
    ambiguous_date_count: int = 0


@dataclass(frozen=True)
class DatasetColumnInspection:
    ordinal: int
    original_header: str
    value_type: str
    valid_count: int
    blank_count: int
    invalid_count: int
    distinct_count: int
    max_label_length: int
    unavailable_reason: str | None = None
    ambiguous_date_count: int = 0


@dataclass(frozen=True)
class DatasetInspection:
    dataset_id: str
    import_spec: DatasetImportSpec
    columns: list[DatasetColumnInspection]
    preview: list[dict[str, str]]
    mapping_version: DatasetMappingVersion | None = None


@dataclass(frozen=True)
class MappingSuggestion:
    column_ordinal: int
    semantic_role: str
    unit: str | None


@dataclass(frozen=True)
class AnalysisEvidence:
    id: int
    thread_id: int
    origin_turn_number: int
    dataset_id: str
    dataset_sha256: str
    import_spec_id: int
    mapping_version_id: int
    operation: str
    schema_version: str


@dataclass(frozen=True)
class ThreadDatasetScope:
    thread_id: int
    dataset_id: str | None
    selected_at: str


class DatasetImportError(ValueError):
    """A local file failed the import safety boundary."""


@dataclass(frozen=True)
class DatasetImportResult:
    dataset: Dataset
    original_path: Path
    duplicate_row_count: int


_SEMANTIC_ROLES = frozenset(
    {"trade_return", "trade_outcome", "trade_timestamp", "session", "direction", "mfe", "mae", "instrument", "setup"}
)
_UNIT_ROLES = frozenset({"trade_return", "mfe", "mae"})
_UNITS = frozenset({"R", "currency", "points", "percentage"})
_DATE_HEADER_TOKENS = frozenset({"date", "time", "timestamp", "datetime"})
_HEADER_ALIASES = {
    "result": ("trade_return", "R"),
    "r": ("trade_return", "R"),
    "resultr": ("trade_return", "R"),
    "pnlr": ("trade_return", "R"),
    "returninr": ("trade_return", "R"),
    "outcome": ("trade_outcome", None),
    "winloss": ("trade_outcome", None),
    "date": ("trade_timestamp", None),
    "time": ("trade_timestamp", None),
    "timestamp": ("trade_timestamp", None),
    "datetime": ("trade_timestamp", None),
    "tradedate": ("trade_timestamp", None),
    "session": ("session", None),
    "direction": ("direction", None),
    "side": ("direction", None),
    "mfe": ("mfe", None),
    "mae": ("mae", None),
    "instrument": ("instrument", None),
    "symbol": ("instrument", None),
    "ticker": ("instrument", None),
    "setup": ("setup", None),
}
_NUMERIC_HEADER_KEYS = frozenset(key for key, (role, _) in _HEADER_ALIASES.items() if role in _UNIT_ROLES)
_OUTCOME_VALUES = {
    "win": "win",
    "w": "win",
    "loss": "loss",
    "l": "loss",
    "breakeven": "breakeven",
    "break even": "breakeven",
    "be": "breakeven",
}


def inspect_local_dataset(storage: "Storage", dataset_id: str, *, preview_rows: int = MAX_INSPECTION_PREVIEW_ROWS) -> DatasetInspection:
    """Read one immutable local original for a bounded, local-only inspection."""
    if not isinstance(preview_rows, int) or not 1 <= preview_rows <= MAX_INSPECTION_PREVIEW_ROWS:
        raise ValueError(f"preview rows must be between 1 and {MAX_INSPECTION_PREVIEW_ROWS}")
    dataset = storage.dataset(dataset_id)
    spec = storage.dataset_import_spec(dataset_id)
    columns = storage.dataset_columns(dataset_id)
    if dataset is None or spec is None or not columns:
        raise ValueError("dataset metadata is unavailable")
    path = storage.database_path.parent / "datasets" / dataset.id / f"original{dataset.original_extension}"
    if not path.is_file():
        raise DatasetImportError("local dataset original is unavailable")
    if _file_sha256(path) != dataset.content_sha256:
        raise DatasetImportError("local dataset original no longer matches its immutable hash")
    headers, rows = _inspection_rows(path, dataset.original_extension, spec)
    expected_headers = [column.original_header for column in columns]
    if headers != expected_headers:
        raise DatasetImportError("local dataset original no longer matches its recorded schema")
    if len(rows) != dataset.source_row_count:
        raise DatasetImportError("local dataset original no longer matches its recorded row count")
    inspected_columns = [_inspect_column(column.ordinal, header, [row[column.ordinal] for row in rows]) for column, header in zip(columns, headers, strict=True)]
    preview = [
        {header: _preview_value(value) for header, value in zip(headers, row, strict=True)}
        for row in rows[:preview_rows]
    ]
    return DatasetInspection(dataset.id, spec, inspected_columns, preview)


def mapping_suggestions(inspection: DatasetInspection) -> list[MappingSuggestion]:
    """Return deterministic header hints; callers must still explicitly confirm them."""
    return [
        MappingSuggestion(column.ordinal, *_HEADER_ALIASES[_header_key(column.original_header)])
        for column in inspection.columns
        if _header_key(column.original_header) in _HEADER_ALIASES
    ]


def create_inspected_mapping_draft(
    storage: "Storage", inspection: DatasetInspection, entries: list[MappingEntry]
) -> DatasetMappingVersion:
    """Persist a local draft with the inspection health snapshot, never active semantics."""
    verified = inspect_local_dataset(storage, inspection.dataset_id)
    if inspection != verified:
        raise ValueError("mapping requires a current local inspection")
    inspection = verified
    columns = _mapping_columns_for_entries(storage, inspection, entries)
    roles: set[str] = set()
    ordinals: set[int] = set()
    prepared: list[MappingEntry] = []
    for entry in entries:
        if not isinstance(entry, MappingEntry) or entry.column_ordinal not in columns:
            raise ValueError("mapping entries must reference an inspected column")
        if entry.column_ordinal in ordinals:
            raise ValueError("mapping columns must be unique")
        ordinals.add(entry.column_ordinal)
        column = columns[entry.column_ordinal]
        if entry.semantic_role is not None:
            if entry.semantic_role not in _SEMANTIC_ROLES:
                raise ValueError("semantic role is unsupported")
            if entry.semantic_role in roles:
                raise ValueError("semantic roles must be unique")
            roles.add(entry.semantic_role)
            allowed_types = {"trade_return": {"number"}, "mfe": {"number"}, "mae": {"number"}, "trade_timestamp": {"datetime"}}.get(
                entry.semantic_role, {"categorical"}
            )
            if column.valid_count and column.value_type not in allowed_types:
                raise ValueError("semantic role is incompatible with the inspected column type")
        if entry.semantic_role in _UNIT_ROLES:
            if entry.unit not in _UNITS:
                raise ValueError("return, MFE, and MAE mappings require a declared unit")
        elif entry.unit is not None:
            raise ValueError("only return, MFE, and MAE mappings may declare a unit")
        if entry.source not in {"manual", "alias"}:
            raise ValueError("mapping source must be manual or alias")
        label = _analysis_label(entry.analysis_label)
        if entry.model_disclosure:
            if label is None:
                raise ValueError("model-disclosed fields require an analysis-safe label")
            if column.value_type not in {"categorical", "boolean"}:
                raise ValueError("aggregate labels may be disclosed only for categorical or boolean fields")
            if column.distinct_count > MAX_MODEL_GROUP_LABELS:
                raise ValueError("aggregate labels may contain at most 20 distinct values")
            if column.max_label_length > MAX_MODEL_GROUP_LABEL_CHARS:
                raise ValueError("aggregate labels may be at most 80 characters")
        if entry.mentor_access not in MENTOR_ACCESS_POLICIES:
            raise ValueError("Mentor access policy is unsupported")
        prepared.append(
            MappingEntry(
                column_ordinal=entry.column_ordinal,
                semantic_role=entry.semantic_role,
                unit=entry.unit,
                analysis_label=label,
                source=entry.source,
                field_id=_field_id(inspection.dataset_id, entry.column_ordinal),
                value_type=column.value_type,
                valid_count=column.valid_count,
                blank_count=column.blank_count,
                invalid_count=column.invalid_count,
                distinct_count=column.distinct_count,
                max_label_length=column.max_label_length,
                aggregate_labels_allowed=entry.model_disclosure,
                mentor_access=entry.mentor_access,
                unavailable_reason=column.unavailable_reason,
                ambiguous_date_count=column.ambiguous_date_count,
            )
        )
    return storage.create_mapping_draft(inspection.dataset_id, prepared)


def model_mapping_context(storage: "Storage", mapping_version_id: int) -> list[dict[str, object]]:
    """Return the confirmed, privacy-safe model contract without raw headers or values."""
    mapping = storage.mapping_version(mapping_version_id)
    if mapping is None or mapping.status != "confirmed":
        return []
    return [
        {
            "field_id": entry.field_id,
            "label": entry.analysis_label if entry.semantic_role is None or entry.aggregate_labels_allowed else None,
            "value_type": entry.value_type,
            "semantic_role": entry.semantic_role,
            "unit": entry.unit,
            "health": {
                "valid_count": entry.valid_count,
                "blank_count": entry.blank_count,
                "invalid_count": entry.invalid_count,
                "ambiguous_date_count": entry.ambiguous_date_count,
                "unavailable_reason": entry.unavailable_reason,
            },
            "aggregate_labels_allowed": entry.aggregate_labels_allowed,
            "mentor_access": entry.mentor_access,
        }
        for entry in storage.mapping_entries(mapping_version_id)
        if entry.field_id is not None
        and (entry.semantic_role is not None or entry.analysis_label is not None)
    ]


def _mapping_columns_for_entries(
    storage: "Storage", inspection: DatasetInspection, entries: Iterable[MappingEntry | object]
) -> dict[int, DatasetColumnInspection]:
    columns = {column.ordinal: column for column in inspection.columns}
    outcome_entries = [entry for entry in entries if _entry_role(entry) == "trade_outcome"]
    if not outcome_entries:
        return columns
    dataset = storage.dataset(inspection.dataset_id)
    if dataset is None:
        raise ValueError("dataset metadata is unavailable")
    headers, rows = _inspection_rows(
        storage.database_path.parent / "datasets" / dataset.id / f"original{dataset.original_extension}",
        dataset.original_extension,
        inspection.import_spec,
    )
    for entry in outcome_entries:
        ordinal = _entry_ordinal(entry)
        if ordinal in columns:
            columns[ordinal] = _inspect_column(
                ordinal, headers[ordinal], [row[ordinal] for row in rows], semantic_role="trade_outcome"
            )
    return columns


def _entry_role(entry: MappingEntry | object) -> object:
    return entry.semantic_role if isinstance(entry, MappingEntry) else getattr(entry, "get", lambda _key: None)("semantic_role")


def _entry_ordinal(entry: MappingEntry | object) -> int | None:
    return entry.column_ordinal if isinstance(entry, MappingEntry) else getattr(entry, "get", lambda _key: None)("column_ordinal")


def _inspection_rows(path: Path, extension: str, spec: DatasetImportSpec) -> tuple[list[str], list[list[object]]]:
    try:
        contents = path.read_bytes()
    except OSError as error:
        raise DatasetImportError("local dataset original is unavailable") from error
    return _inspection_rows_from_bytes(contents, extension, spec)


def _inspection_rows_from_bytes(
    contents: bytes, extension: str, spec: DatasetImportSpec
) -> tuple[list[str], list[list[object]]]:
    if extension == ".csv":
        if not spec.csv_encoding or not spec.csv_delimiter or not spec.csv_quoting:
            raise DatasetImportError("CSV import specification is incomplete")
        try:
            text = contents.decode(spec.csv_encoding, errors="strict")
        except UnicodeDecodeError as error:
            raise DatasetImportError("CSV parse failed") from error
        rows = csv.reader(io.StringIO(text, newline=""), delimiter=spec.csv_delimiter, quotechar=spec.csv_quoting, strict=True)
        try:
            headers = _headers(next(rows))
        except StopIteration as error:
            raise DatasetImportError("CSV has no header row") from error
        values = [list(row) for row in rows if row and not all(value == "" for value in row)]
    else:
        _preflight_xlsx_bytes(contents, spec.selected_sheet)
        try:
            workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=False, keep_links=False)
        except Exception as error:
            raise DatasetImportError("XLSX parse failed") from error
        try:
            worksheet = workbook[spec.selected_sheet or ""]
            worksheet.reset_dimensions()
            iterator = worksheet.iter_rows(values_only=True)
            headers = _headers(next(iterator))
            values = [list(row) for row in iterator if row and not all(value is None or value == "" for value in row)]
        except KeyError as error:
            raise DatasetImportError("selected XLSX worksheet does not exist") from error
        except StopIteration as error:
            raise DatasetImportError("XLSX has no header row") from error
        finally:
            workbook.close()
    if any(len(row) != len(headers) for row in values):
        raise DatasetImportError("dataset rows do not match the header")
    return headers, values


def _inspect_column(
    ordinal: int, header: str, values: list[object], *, semantic_role: str | None = None
) -> DatasetColumnInspection:
    nonblank = [value for value in values if not _blank(value)]
    blank_count = len(values) - len(nonblank)
    text = [str(value).strip() for value in nonblank]
    if not text:
        return DatasetColumnInspection(ordinal, header, "unknown", 0, blank_count, 0, 0, 0, "no_valid_values")
    if semantic_role == "trade_outcome":
        return _outcome_column_inspection(ordinal, header, text, blank_count)
    if _is_date_header(header):
        date_states = [_date_parse_state(value) for value in text]
        valid_dates = date_states.count("valid")
        invalid_count = len(text) - valid_dates
        ambiguous_date_count = date_states.count("ambiguous")
        reason = "ambiguous_date" if ambiguous_date_count else "invalid_values_excluded" if invalid_count else None
        return DatasetColumnInspection(
            ordinal,
            header,
            "datetime",
            valid_dates,
            blank_count,
            invalid_count,
            len(set(text)),
            max(map(len, text)),
            reason,
            ambiguous_date_count,
        )
    if _HEADER_ALIASES.get(_header_key(header), (None,))[0] == "trade_outcome":
        return _outcome_column_inspection(ordinal, header, text, blank_count)
    numeric_count = sum(_parse_number(value) is not None for value in text)
    if numeric_count and (numeric_count == len(text) or _header_key(header) in _NUMERIC_HEADER_KEYS):
        invalid_count = len(text) - numeric_count
        return DatasetColumnInspection(
            ordinal,
            header,
            "number",
            numeric_count,
            blank_count,
            invalid_count,
            len(set(text)),
            max(map(len, text)),
            "invalid_values_excluded" if invalid_count else None,
        )
    normalized = {value.casefold() for value in text}
    if normalized.issubset({"true", "false", "yes", "no", "y", "n"}):
        return DatasetColumnInspection(ordinal, header, "boolean", len(text), blank_count, 0, len(normalized), max(map(len, text)))
    return DatasetColumnInspection(ordinal, header, "categorical", len(text), blank_count, 0, len(set(text)), max(map(len, text)))


def _outcome_column_inspection(
    ordinal: int, header: str, text: list[str], blank_count: int
) -> DatasetColumnInspection:
    normalized_outcomes = [_OUTCOME_VALUES.get(value.casefold()) for value in text]
    valid_count = sum(value is not None for value in normalized_outcomes)
    invalid_count = len(text) - valid_count
    return DatasetColumnInspection(
        ordinal,
        header,
        "categorical",
        valid_count,
        blank_count,
        invalid_count,
        len({value for value in normalized_outcomes if value is not None}),
        max((len(value) for value in normalized_outcomes if value is not None), default=0),
        "invalid_values_excluded" if invalid_count else None,
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        while chunk := file.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _is_date_header(header: str) -> bool:
    return _HEADER_ALIASES.get(_header_key(header), (None,))[0] == "trade_timestamp" or bool(
        _DATE_HEADER_TOKENS.intersection(re.findall(r"[a-z]+", header.casefold()))
    )


def _parse_number(value: str) -> float | None:
    try:
        parsed = float(value[:-1]) / 100 if value.endswith("%") else float(value)
        return parsed if math.isfinite(parsed) else None
    except ValueError:
        return None


def _date_parse_state(value: str) -> str:
    if "/" in value:
        return "ambiguous"
    try:
        datetime.fromisoformat(value.replace("Z", "+00:00"))
        return "valid"
    except ValueError:
        return "invalid"


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _preview_value(value: object) -> str:
    text = "" if value is None else str(value)
    return text[:MAX_INSPECTION_CELL_CHARS]


def _analysis_label(value: str | None) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or not (label := " ".join(value.split())) or len(label) > 80:
        raise ValueError("analysis-safe label must be 1 to 80 characters")
    return label


def _field_id(dataset_id: str, ordinal: int) -> str:
    return f"field_{hashlib.sha256(f'{dataset_id}:{ordinal}'.encode()).hexdigest()[:12]}"


def import_local_dataset(
    source_path: Path,
    storage: "Storage",
    datasets_root: Path | None = None,
    *,
    selected_sheet: str | None = None,
    dataset_id_factory: Callable[[], str] | None = None,
) -> DatasetImportResult:
    """Copy, validate, and record one local CSV/XLSX upload without retaining rows."""
    source_path = Path(source_path)
    extension = source_path.suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        raise DatasetImportError("only CSV or XLSX files are supported")
    if not source_path.is_file():
        raise DatasetImportError("dataset file does not exist")
    if source_path.stat().st_size > MAX_DATASET_BYTES:
        raise DatasetImportError("dataset exceeds the 50 MiB size limit")

    datasets_root = _internal_datasets_root(storage, datasets_root)
    with _dataset_import_lease(datasets_root):
        return _import_local_dataset_locked(
            source_path,
            storage,
            datasets_root,
            extension=extension,
            selected_sheet=selected_sheet,
            dataset_id_factory=dataset_id_factory,
        )


def _import_local_dataset_locked(
    source_path: Path,
    storage: "Storage",
    datasets_root: Path,
    *,
    extension: str,
    selected_sheet: str | None,
    dataset_id_factory: Callable[[], str] | None,
) -> DatasetImportResult:
    _recover_pending_dataset_imports(datasets_root, storage)
    temporary_directory = Path(tempfile.mkdtemp(prefix=".import-", dir=datasets_root))
    final_directory: Path | None = None
    dataset_id: str | None = None
    reserved_this_import = False
    try:
        temporary_original = temporary_directory / f"original{extension}"
        content_sha256, byte_size = _copy_with_hash(source_path, temporary_original)
        if extension == ".csv":
            parsed = _parse_csv(temporary_original)
        else:
            parsed = _parse_xlsx(temporary_original, selected_sheet)

        dataset_id = (dataset_id_factory or _new_dataset_id)()
        if not isinstance(dataset_id, str) or _DATASET_ID_PATTERN.fullmatch(dataset_id) is None:
            raise DatasetImportError("dataset identifier is invalid")
        candidate_directory = datasets_root / dataset_id
        if candidate_directory.exists():
            raise DatasetImportError("dataset identifier already exists")
        try:
            storage.begin_dataset_import(dataset_id)
        except ValueError as error:
            raise DatasetImportError("dataset identifier already exists") from error
        reserved_this_import = True
        (temporary_directory / ".pending-import").touch(exist_ok=False)
        final_directory = candidate_directory
        temporary_directory.replace(final_directory)
        original_path = final_directory / temporary_original.name
        dataset = storage.create_dataset(
            dataset_id=dataset_id,
            original_name=source_path.name,
            content_sha256=content_sha256,
            original_extension=extension,
            byte_size=byte_size,
            source_row_count=parsed.row_count,
            status="ready",
            import_spec=DatasetImportSpec(
                header_row=0,
                selected_sheet=parsed.selected_sheet,
                csv_encoding=parsed.csv_encoding,
                csv_delimiter=parsed.csv_delimiter,
                csv_quoting=parsed.csv_quoting,
                parser_version=parsed.parser_version,
                row_order_policy="source",
                time_parse_policy="unambiguous_only",
            ),
            columns=[
                DatasetColumn(index, header, "unknown", null_count, 0)
                for index, (header, null_count) in enumerate(zip(parsed.headers, parsed.null_counts, strict=True))
            ],
        )
        storage.mark_dataset_import_committed(dataset_id)
        (final_directory / ".pending-import").unlink()
        storage.complete_dataset_import(dataset_id)
        return DatasetImportResult(dataset, original_path, parsed.duplicate_row_count)
    except BaseException:
        if reserved_this_import and dataset_id is not None and storage.dataset_import_state(dataset_id) == "staging":
            try:
                if final_directory is not None and (final_directory / ".pending-import").exists():
                    shutil.rmtree(final_directory)
                storage.discard_failed_dataset_import(dataset_id)
                storage.abandon_dataset_import(dataset_id)
            except Exception:
                pass
        raise
    finally:
        if temporary_directory.exists():
            shutil.rmtree(temporary_directory)


@contextmanager
def _dataset_import_lease(datasets_root: Path) -> Iterator[None]:
    """Serialize local import recovery and promotion across threads and processes."""
    with _IMPORT_LEASE_LOCK:
        lease_path = datasets_root.parent / ".dataset-import.lock"
        with lease_path.open("a+b") as lease_file:
            lease_file.seek(0, 2)
            if lease_file.tell() == 0:
                lease_file.write(b"\0")
                lease_file.flush()
            lease_file.seek(0)
            _acquire_file_lease(lease_file)
            try:
                yield
            finally:
                _release_file_lease(lease_file)


def _acquire_file_lease(lease_file: io.BufferedRandom) -> None:
    if os.name == "nt":
        import msvcrt

        while True:
            try:
                msvcrt.locking(lease_file.fileno(), msvcrt.LK_NBLCK, 1)
                return
            except OSError:
                time.sleep(0.05)
    else:
        import fcntl

        fcntl.flock(lease_file.fileno(), fcntl.LOCK_EX)


def _release_file_lease(lease_file: io.BufferedRandom) -> None:
    lease_file.seek(0)
    if os.name == "nt":
        import msvcrt

        msvcrt.locking(lease_file.fileno(), msvcrt.LK_UNLCK, 1)
    else:
        import fcntl

        fcntl.flock(lease_file.fileno(), fcntl.LOCK_UN)


def _internal_datasets_root(storage: "Storage", requested_root: Path | None) -> Path:
    data_root = storage.database_path.parent.resolve(strict=False)
    expected_root = data_root / "datasets"
    if str(data_root).startswith("\\\\") or str(expected_root).startswith("\\\\"):
        raise DatasetImportError("dataset root must be local")
    try:
        resolved_root = expected_root.resolve(strict=False)
        resolved_root.relative_to(data_root)
    except (OSError, ValueError) as error:
        raise DatasetImportError("dataset root escapes local data storage") from error
    if requested_root is not None:
        try:
            requested_root = Path(requested_root)
            if str(requested_root).startswith("\\\\") or requested_root.resolve(strict=False) != resolved_root:
                raise DatasetImportError("dataset root is internal and cannot be caller controlled")
        except OSError as error:
            raise DatasetImportError("dataset root is invalid") from error
    resolved_root.mkdir(parents=True, exist_ok=True)
    return resolved_root


def _recover_pending_dataset_imports(datasets_root: Path, storage: "Storage") -> None:
    for dataset_id in storage.pending_dataset_import_ids():
        pending_directory = datasets_root / dataset_id
        if storage.dataset_import_state(dataset_id) == "staging":
            if pending_directory.exists():
                shutil.rmtree(pending_directory)
            storage.discard_failed_dataset_import(dataset_id)
            storage.abandon_dataset_import(dataset_id)
        else:
            marker = pending_directory / ".pending-import"
            if marker.exists():
                marker.unlink()
            storage.complete_dataset_import(dataset_id)
    for temporary_directory in datasets_root.glob(".import-*"):
        if temporary_directory.is_dir():
            shutil.rmtree(temporary_directory)


@dataclass(frozen=True)
class _ParsedDataset:
    headers: list[str]
    null_counts: list[int]
    row_count: int
    duplicate_row_count: int
    selected_sheet: str | None
    csv_encoding: str | None
    csv_delimiter: str | None
    csv_quoting: str | None
    parser_version: str


def _new_dataset_id() -> str:
    return f"dataset-{uuid.uuid4().hex}"


def _copy_with_hash(source_path: Path, destination_path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    byte_size = 0
    with source_path.open("rb") as source, destination_path.open("xb") as destination:
        while chunk := source.read(1024 * 1024):
            byte_size += len(chunk)
            if byte_size > MAX_DATASET_BYTES:
                raise DatasetImportError("dataset exceeds the 50 MiB size limit")
            digest.update(chunk)
            destination.write(chunk)
    return digest.hexdigest(), byte_size


def _parse_csv(path: Path) -> _ParsedDataset:
    raw = path.read_bytes()
    encoding = _csv_encoding(raw)
    try:
        text = raw.decode(encoding, errors="strict")
    except UnicodeDecodeError as error:
        raise DatasetImportError("CSV decode failed") from error
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error as error:
        if any(delimiter in text for delimiter in ";\t|"):
            raise DatasetImportError("CSV dialect could not be determined") from error
        dialect = csv.excel
    if dialect.delimiter not in {",", ";", "\t", "|"} or dialect.quotechar != '"':
        raise DatasetImportError("CSV dialect is unsupported")
    try:
        rows = csv.reader(io.StringIO(text, newline=""), dialect=dialect, strict=True)
        headers = _headers(next(rows))
        return _parsed_rows(
            headers,
            rows,
            selected_sheet=None,
            csv_encoding=encoding,
            csv_delimiter=dialect.delimiter,
            csv_quoting=dialect.quotechar,
            parser_version="csv-stdlib-1",
        )
    except StopIteration as error:
        raise DatasetImportError("CSV has no header row") from error
    except csv.Error as error:
        raise DatasetImportError("CSV parse failed") from error


def _csv_encoding(raw: bytes) -> str:
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    return "utf-8-sig" if raw.startswith(b"\xef\xbb\xbf") else "utf-8"


def _parse_xlsx(path: Path, selected_sheet: str | None) -> _ParsedDataset:
    selected_sheet = _preflight_xlsx(path, selected_sheet)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception as error:
        raise DatasetImportError("XLSX parse failed") from error
    try:
        worksheet_names = [worksheet.title for worksheet in workbook.worksheets]
        if selected_sheet not in worksheet_names:
            raise DatasetImportError("selected XLSX worksheet does not exist")
        worksheet = workbook[selected_sheet]
        worksheet.reset_dimensions()
        values = worksheet.iter_rows(values_only=True)
        try:
            headers = _headers(next(values))
        except StopIteration as error:
            raise DatasetImportError("XLSX has no header row") from error
        return _parsed_rows(
            headers,
            values,
            selected_sheet=selected_sheet,
            csv_encoding=None,
            csv_delimiter=None,
            csv_quoting=None,
            parser_version="openpyxl-3.1",
        )
    finally:
        workbook.close()


def _preflight_xlsx(path: Path, selected_sheet: str | None) -> str:
    try:
        contents = path.read_bytes()
    except OSError as error:
        raise DatasetImportError("XLSX signature is invalid") from error
    return _preflight_xlsx_bytes(contents, selected_sheet)


def _preflight_xlsx_bytes(contents: bytes, selected_sheet: str | None) -> str:
    if contents[:4] != b"PK\x03\x04" or not zipfile.is_zipfile(io.BytesIO(contents)):
        raise DatasetImportError("XLSX signature is invalid")
    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as archive:
            members = archive.infolist()
            names = [member.filename for member in members]
            compressed_bytes = sum(member.compress_size for member in members)
            uncompressed_bytes = sum(member.file_size for member in members)
            if (
                len(members) > MAX_XLSX_ARCHIVE_MEMBERS
                or compressed_bytes > MAX_XLSX_ARCHIVE_COMPRESSED_BYTES
                or uncompressed_bytes > MAX_XLSX_ARCHIVE_UNCOMPRESSED_BYTES
                or len(names) != len(set(names))
            ):
                raise DatasetImportError("XLSX archive exceeds safety limits")
            for member in members:
                if (
                    member.flag_bits & 0x1
                    or not _canonical_zip_member_name(member)
                    or (member.compress_size and member.file_size > member.compress_size * MAX_XLSX_COMPRESSION_RATIO)
                ):
                    raise DatasetImportError("XLSX archive is unsafe")
            if not {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}.issubset(names):
                raise DatasetImportError("XLSX OOXML signature is invalid")
            _reject_macro_content_types(archive.read("[Content_Types].xml"))
            relationships = {
                name: _parse_relationships(archive.read(name))
                for name in names
                if name.endswith(".rels")
            }
            if any("vbaproject" in name.casefold() for name in names) or any(
                name.startswith("xl/externalLinks/") for name in names
            ):
                raise DatasetImportError("XLSX macros or external links are not supported")
            sheets = _workbook_sheets(
                archive.read("xl/workbook.xml"), relationships.get("xl/_rels/workbook.xml.rels", {})
            )
            selected_sheet = selected_sheet or next(iter(sheets), None)
            if selected_sheet not in sheets:
                raise DatasetImportError("selected XLSX worksheet does not exist")
            for name, worksheet_path in sheets.items():
                _scan_worksheet_xml(archive.read(worksheet_path), enforce_limits=name == selected_sheet)
            return selected_sheet
    except zipfile.BadZipFile as error:
        raise DatasetImportError("XLSX archive is invalid") from error


def _canonical_zip_member_name(member: zipfile.ZipInfo) -> bool:
    name = member.filename
    canonical = name[:-1] if member.is_dir() else name
    return bool(
        canonical
        and "\\" not in canonical
        and not canonical.startswith("/")
        and re.match(r"^[A-Za-z]:", canonical) is None
        and all(part not in {"", ".", ".."} for part in canonical.split("/"))
    )


def _xml_document(contents: bytes, label: str) -> ElementTree.Element:
    try:
        return ElementTree.fromstring(contents)
    except ElementTree.ParseError as error:
        raise DatasetImportError(f"XLSX {label} XML is invalid") from error


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _reject_macro_content_types(contents: bytes) -> None:
    document = _xml_document(contents, "content types")
    for element in document.iter():
        content_type = element.attrib.get("ContentType", "").casefold()
        if "macroenabled" in content_type or "vba" in content_type:
            raise DatasetImportError("XLSX macros are not supported")


def _parse_relationships(contents: bytes) -> dict[str, str]:
    document = _xml_document(contents, "relationship")
    relationships: dict[str, str] = {}
    for element in document:
        if _local_name(element.tag) != "Relationship":
            continue
        relationship_id = element.attrib.get("Id")
        relationship_type = element.attrib.get("Type", "").casefold()
        target = element.attrib.get("Target", "")
        target_mode = element.attrib.get("TargetMode", "").casefold()
        normalized_target = target.strip()
        if (
            target_mode == "external"
            or "externallink" in relationship_type
            or normalized_target != target
            or normalized_target.startswith("//")
            or "\\" in normalized_target
            or re.match(r"^[A-Za-z][A-Za-z0-9+.-]*:", normalized_target) is not None
        ):
            raise DatasetImportError("XLSX external links are not supported")
        if "vbaproject" in relationship_type or target.casefold().endswith("vbaproject.bin"):
            raise DatasetImportError("XLSX macros are not supported")
        if not relationship_id or not target:
            raise DatasetImportError("XLSX relationship is invalid")
        if relationship_id in relationships:
            raise DatasetImportError("XLSX relationship identifiers must be unique")
        relationships[relationship_id] = target
    return relationships


def _workbook_sheets(workbook_contents: bytes, relationships: dict[str, str]) -> dict[str, str]:
    document = _xml_document(workbook_contents, "workbook")
    sheets: dict[str, str] = {}
    for element in document.iter():
        if _local_name(element.tag) != "sheet":
            continue
        name = element.attrib.get("name")
        relationship_id = next((value for key, value in element.attrib.items() if _local_name(key) == "id"), None)
        target = relationships.get(relationship_id or "")
        if not name or not target or name in sheets:
            raise DatasetImportError("XLSX workbook sheet relationship is invalid")
        sheets[name] = _resolve_package_target("xl/workbook.xml", target)
    if not sheets:
        raise DatasetImportError("XLSX has no worksheets")
    return sheets


def _resolve_package_target(owner: str, target: str) -> str:
    if "\\" in target or re.match(r"^[A-Za-z]:", target):
        raise DatasetImportError("XLSX relationship target is invalid")
    if target.startswith("/"):
        if not target.startswith("/xl/"):
            raise DatasetImportError("XLSX relationship target is invalid")
        parts: list[str] = []
        target = target[1:]
    else:
        parts = list(PurePosixPath(owner).parent.parts)
    for part in target.split("/"):
        if part in {"", "."}:
            raise DatasetImportError("XLSX relationship target is invalid")
        if part == "..":
            if not parts:
                raise DatasetImportError("XLSX relationship target is invalid")
            parts.pop()
        else:
            parts.append(part)
    return "/".join(parts)


def _scan_worksheet_xml(contents: bytes, *, enforce_limits: bool) -> None:
    cell_count = 0
    try:
        for event, element in ElementTree.iterparse(io.BytesIO(contents), events=("end",)):
            name = _local_name(element.tag)
            if name == "f":
                raise DatasetImportError("XLSX formula cells are not supported")
            if name == "c" and enforce_limits:
                reference = element.attrib.get("r", "")
                match = _CELL_REFERENCE_PATTERN.fullmatch(reference)
                if match is None:
                    raise DatasetImportError("XLSX cell reference is invalid")
                column = _column_number(match.group(1))
                row = int(match.group(2))
                cell_count += 1
                if (
                    column > MAX_DATASET_COLUMNS
                    or row > MAX_DATASET_ROWS + 1
                    or cell_count > MAX_DATASET_CELLS + MAX_DATASET_COLUMNS
                ):
                    raise DatasetImportError("XLSX exceeds row, column, or cell limits")
            element.clear()
    except ElementTree.ParseError as error:
        raise DatasetImportError("XLSX worksheet XML is invalid") from error


def _column_number(label: str) -> int:
    value = 0
    for character in label:
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _headers(row: Iterable[object]) -> list[str]:
    headers = list(row)
    if not headers:
        raise DatasetImportError("dataset has no header row")
    if len(headers) > MAX_DATASET_COLUMNS:
        raise DatasetImportError("dataset exceeds the column limit")
    if any(not isinstance(header, str) or not header.strip() for header in headers):
        raise DatasetImportError("dataset headers must be non-empty text")
    if len(set(headers)) != len(headers):
        raise DatasetImportError("dataset headers must be unique")
    return headers


def _parsed_rows(
    headers: list[str],
    rows: Iterable[Iterable[object]],
    *,
    selected_sheet: str | None,
    csv_encoding: str | None,
    csv_delimiter: str | None,
    csv_quoting: str | None,
    parser_version: str,
) -> _ParsedDataset:
    row_count = 0
    duplicate_row_count = 0
    null_counts = [0] * len(headers)
    row_hashes: set[bytes] = set()
    for row in rows:
        values = list(row)
        if not values or all(value is None or value == "" for value in values):
            continue
        if len(values) != len(headers):
            raise DatasetImportError("dataset rows do not match the header")
        row_count += 1
        if row_count > MAX_DATASET_ROWS:
            raise DatasetImportError("dataset exceeds the row limit")
        if row_count * len(headers) > MAX_DATASET_CELLS:
            raise DatasetImportError("dataset exceeds the cell limit")
        for index, value in enumerate(values):
            if value is None or value == "":
                null_counts[index] += 1
        row_hash = _row_hash(values)
        if row_hash in row_hashes:
            duplicate_row_count += 1
        else:
            row_hashes.add(row_hash)
    return _ParsedDataset(
        headers,
        null_counts,
        row_count,
        duplicate_row_count,
        selected_sheet,
        csv_encoding,
        csv_delimiter,
        csv_quoting,
        parser_version,
    )


def _row_hash(values: list[object]) -> bytes:
    digest = hashlib.sha256()
    for value in values:
        encoded = f"{type(value).__name__}:{value!r}".encode("utf-8")
        digest.update(len(encoded).to_bytes(8, "big"))
        digest.update(encoded)
    return digest.digest()
